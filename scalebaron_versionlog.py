# Scalebaron - Clean Rotation + Version Tracking
# ------------------------------------------------
# Adds:
# - Preprocessing inventory
# - Optional image rotation
# - Layout choice (portrait or landscape)
# - Improved sample labeling
# - Version log auto-updating

import os
import re
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib import cm
from PIL import Image, ImageDraw, ImageFont
from scipy import stats

# === Settings ===
VERSION_NAME = "scalebaron_clean_rotation"
PIXEL_SIZE = float(input("Enter pixel size in microns: "))
SCALE_BAR_LENGTH_UM = 1000
SCALE_BAR_PIXELS = int(SCALE_BAR_LENGTH_UM // PIXEL_SIZE)
INPUT_DIR = "./INPUT"
OUTPUT_DIR = "./OUTPUT"
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

color_scheme = input("Enter a Matplotlib color scheme (e.g., jet, viridis): ") or "jet"
auto_accept = input("Auto-accept all suggested scale max values? (y/n): ").strip().lower() == 'y'

BACKGROUND_COLOR = tuple([int(c * 255) for c in plt.get_cmap(color_scheme)(0)[:3]])

# === Functions ===
def load_matrix_2d(path):
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    return np.array([[cell if isinstance(cell, (int, float)) else np.nan for cell in row] for row in ws.iter_rows(values_only=True)])

def compute_99th_percentile(values):
    return np.nanpercentile(values.flatten(), 99)

def compute_variance(values):
    return np.nanvar(values.flatten())

def plot_matrix(matrix, scale_max, output_path, rotate=False):
    if rotate and matrix.shape[0] > matrix.shape[1]:
        matrix = np.rot90(matrix)
    fig, ax = plt.subplots(figsize=(4, 4), facecolor=plt.get_cmap(color_scheme)(0))
    cmap = plt.get_cmap(color_scheme)
    norm = mcolors.Normalize(vmin=0, vmax=scale_max)
    ax.imshow(matrix, cmap=cmap, norm=norm)
    ax.axis('off')
    bar_x = matrix.shape[1] - SCALE_BAR_PIXELS - 10
    bar_y = matrix.shape[0] - 10
    ax.hlines(bar_y, bar_x, bar_x + SCALE_BAR_PIXELS, colors='white', linewidth=2)
    plt.tight_layout(pad=0)
    fig.savefig(output_path, dpi=300, bbox_inches='tight', pad_inches=0)
    plt.close(fig)
    return output_path

def label_image(img_path, label_text, font_size=72):
    img = Image.open(img_path)
    new_height = img.height + font_size + 10
    labeled = Image.new("RGB", (img.width, new_height), BACKGROUND_COLOR)
    labeled.paste(img, (0, 0))
    draw = ImageDraw.Draw(labeled)
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except:
        try:
            font = ImageFont.truetype("DejaVuSans.ttf", font_size)
        except:
            font = ImageFont.load_default()
    text_width = draw.textlength(label_text, font=font)
    draw.text(((img.width - text_width) / 2, img.height + 5), label_text, fill="white", font=font)
    labeled_path = img_path.replace(".png", "_labeled.png")
    labeled.save(labeled_path)
    return labeled_path

def save_colorbar_horizontal(scale_max, path, unit):
    fig, ax = plt.subplots(figsize=(4, 0.4))
    cmap = plt.get_cmap(color_scheme)
    norm = mcolors.Normalize(vmin=0, vmax=scale_max)
    cb = cm.ScalarMappable(norm=norm, cmap=cmap)
    cbar = plt.colorbar(cb, cax=ax, orientation='horizontal')
    cbar.set_label(unit, fontsize=12)
    cbar.ax.tick_params(labelsize=10)
    fig.savefig(path, dpi=300, bbox_inches='tight', transparent=True)
    plt.close(fig)

def resize_final_image(path, max_width=2400, max_height=1800):
    img = Image.open(path)
    img.thumbnail((max_width, max_height), Image.LANCZOS)
    img.save(path)

def create_composite(image_paths, colorbar_path, out_path, rows, cols, layout_orientation):
    images = [Image.open(p) for p in image_paths]
    max_w = max(im.width for im in images)
    max_h = max(im.height for im in images)
    padding = 10
    colorbar = Image.open(colorbar_path)

    if layout_orientation == "landscape":
        rows, cols = cols, rows

    total_w = cols * max_w + (cols + 1) * padding
    total_h = rows * max_h + (rows + 1) * padding + colorbar.height + padding

    composite = Image.new("RGB", (total_w, total_h), BACKGROUND_COLOR)

    for i, img in enumerate(images):
        row = i // cols
        col = i % cols
        x = padding + col * (max_w + padding)
        y = padding + row * (max_h + padding)
        composite.paste(img, (x, y))

    cb_x = (total_w - colorbar.width) // 2
    cb_y = total_h - colorbar.height - padding
    composite.paste(colorbar, (cb_x, cb_y))
    composite.save(out_path)
    resize_final_image(out_path)

def update_version_log():
    log_path = os.path.join(OUTPUT_DIR, "Scalebaron_version_log.txt")
    entry = f"| {pd.Timestamp.now().date()} | {VERSION_NAME} | Added image rotation, layout choice, input inventory, summary table restored |"
    if os.path.exists(log_path):
        with open(log_path, "a") as f:
            f.write(entry)
    else:
        with open(log_path, "w") as f:
            f.write("| Date | Version Name | Major Changes |\n|:------------|:-------------------------|:-------------------------------------------------|")
            f.write(entry)

# === Main Script ===

print("\n🔍 Scanning INPUT folder...")
matrix_files = sorted(f for f in os.listdir(INPUT_DIR) if f.endswith(".xlsx") and not f.startswith("~"))
print(f"Found {len(matrix_files)} matrix files.")

samples_by_element = {}
for file in matrix_files:
    match = re.match(r"(.+?)[_ ]([A-Za-z]{1,2}\d{2,3})_(ppm|CPS) matrix\.xlsx", file)
    if match:
        sample, element, unit_type = match.groups()
        key = f"{element}_{unit_type}"
        samples_by_element.setdefault(key, []).append((sample, os.path.join(INPUT_DIR, file)))
    else:
        print(f"⚠️ Skipped: {file}")

rotate_decision = input("Rotate images where needed for longest side horizontal? (y/n): ").strip().lower() == 'y'
layout_orientation = input("Composite orientation: portrait or landscape? (p/l): ").strip().lower()
layout_orientation = "portrait" if layout_orientation == "p" else "landscape"

summary_rows = []
variance_rows = []

for key, sample_list in samples_by_element.items():
    element, unit_type = key.split("_")
    unit_label = "ppm" if unit_type == "ppm" else "CPS"
    element_output_dir = os.path.join(OUTPUT_DIR, element)
    os.makedirs(element_output_dir, exist_ok=True)

    print(f"\n📊 Processing {len(sample_list)} sample(s) for {element} [{unit_label}]...")

    all_values = []
    percentiles = {}
    variances = {}
    image_paths = []

    for sample, filepath in sample_list:
        matrix = load_matrix_2d(filepath)
        all_values.append(matrix)
        percentiles[sample] = compute_99th_percentile(matrix)
        variances[sample] = compute_variance(matrix)
        print(f"  • {sample}: 99th = {percentiles[sample]:.1f}, var = {variances[sample]:.1f}")

    concat_values = np.concatenate([m.flatten() for m in all_values])
    scale_99 = compute_99th_percentile(concat_values)

    if not auto_accept:
        user_input = input(f"Suggested 99th percentile max: {scale_99:.2f}. Accept? (Enter or new value): ")
        if user_input.strip():
            try:
                scale_99 = float(user_input)
            except ValueError:
                pass

    colorbar_path = os.path.join(element_output_dir, f"{element}_colorbar.png")
    save_colorbar_horizontal(scale_99, colorbar_path, unit_label)

    for sample, filepath in sample_list:
        matrix = load_matrix_2d(filepath)
        img_path = plot_matrix(matrix, scale_99, os.path.join(element_output_dir, f"{sample}_{element}.png"), rotate=rotate_decision)
        labeled_path = label_image(img_path, sample, 72)
        image_paths.append(labeled_path)

    composite_path = os.path.join(element_output_dir, f"{element}_composite.png")

    cols = int(input("Enter number of columns: "))
    rows = int(input("Enter number of rows: "))
    create_composite(image_paths, colorbar_path, composite_path, rows, cols, layout_orientation)

    for sample in percentiles:
        summary_rows.append({"Sample": sample, element: round(percentiles[sample], 1)})
        variance_rows.append({"Sample": sample, element: round(variances[sample], 1)})

# Save summaries
if summary_rows:
    df1 = pd.DataFrame(summary_rows).groupby("Sample").first().reset_index()
    df2 = pd.DataFrame(variance_rows).groupby("Sample").first().reset_index()
    df1.to_csv(os.path.join(OUTPUT_DIR, "summary_99th_percentiles.csv"), index=False)
    df2.to_csv(os.path.join(OUTPUT_DIR, "summary_variances.csv"), index=False)
    print("\n📄 Summary tables saved.")

# Save version log
update_version_log()

print("\n✅ All processing complete. Output in:", OUTPUT_DIR)
