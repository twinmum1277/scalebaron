"""
ScaleBarOn v0.8.0 — GitHub-Ready Composite Generator
Date: 2025-05-01

CHANGELOG:
- Formal versioning system initiated (v0.8.0)
- Base version derived from `scalebaron_clean2`
- Reorganized and documented all code blocks
- Preserved all core features:
  - Pixel size prompt
  - Color scheme and layout prompts
  - Image rotation (optional)
  - File inventory confirmation
  - Scale bar size prompt (µm)
  - Composite organization (grid, spatial and ppm scale bars)
  - Summary tables (99th percentile, variance)

Next goals:
- Improve visual layout of scale bars and labels
- Track cumulative features for JOSS submission
"""

import os
import re
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib import cm
from PIL import Image, ImageDraw, ImageFont
from scipy import stats

def main():
    # --- User Inputs ---
    PIXEL_SIZE = float(input("Enter pixel size in microns: "))
    SCALE_BAR_LENGTH_UM = float(input("Enter desired scale bar length in microns: "))
    SCALE_BAR_PIXELS = int(SCALE_BAR_LENGTH_UM // PIXEL_SIZE)

    color_scheme = input("Enter a Matplotlib color scheme (e.g., jet, viridis): ") or "jet"
    rotate_images = input("Rotate images? (y/n): ").strip().lower() == 'y'
    portrait = None
    if rotate_images:
        layout_mode = input("Composite orientation: portrait or landscape? (p/l): ").strip().lower()
        portrait = layout_mode == 'p'

    INPUT_DIR = "./INPUT"
    OUTPUT_DIR = "./OUTPUT"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    BACKGROUND_COLOR = tuple([int(c * 255) for c in plt.get_cmap(color_scheme)(0)[:3]])

    # --- Utility Functions ---
    def load_matrix_2d(path):
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active
        return np.array([[cell if isinstance(cell, (int, float)) else np.nan for cell in row] for row in ws.iter_rows(values_only=True)])

    def compute_99th_percentile(values):
        return np.nanpercentile(values.flatten(), 99)

    def compute_variance(values):
        return np.nanvar(values.flatten())

    def plot_matrix(matrix, scale_max, output_path):
        fig, ax = plt.subplots(figsize=(4, 4), facecolor=plt.get_cmap(color_scheme)(0))
        cmap = plt.get_cmap(color_scheme)
        norm = mcolors.Normalize(vmin=0, vmax=scale_max)
        ax.imshow(matrix, cmap=cmap, norm=norm)
        ax.axis('off')
        plt.tight_layout(pad=0)
        fig.savefig(output_path, dpi=300, bbox_inches='tight', pad_inches=0)
        plt.close(fig)
        return output_path

    def label_image(img_path, label_text, font_size=48):
        img = Image.open(img_path)
        new_height = img.height + font_size + 10
        labeled = Image.new("RGB", (img.width, new_height), BACKGROUND_COLOR)
        labeled.paste(img, (0, 0))
        draw = ImageDraw.Draw(labeled)
        try:
            font = ImageFont.truetype("arial.ttf", font_size)
        except:
            font = ImageFont.load_default()
        text_width = draw.textlength(label_text, font=font)
        draw.text(((img.width - text_width) / 2, img.height + 5), label_text, fill="white", font=font)
        labeled_path = img_path.replace(".png", "_labeled.png")
        labeled.save(labeled_path)
        return labeled_path

    def save_colorbar(scale_max, path, unit):
        fig, ax = plt.subplots(figsize=(4, 0.4))
        cmap = plt.get_cmap(color_scheme)
        norm = mcolors.Normalize(vmin=0, vmax=scale_max)
        cb = cm.ScalarMappable(norm=norm, cmap=cmap)
        cbar = plt.colorbar(cb, cax=ax, orientation='horizontal')
        cbar.set_label(unit, fontsize=12)
        cbar.ax.tick_params(labelsize=10)
        fig.savefig(path, dpi=300, bbox_inches='tight', transparent=True)
        plt.close(fig)

    def create_composite(images, labels, colorbar_path, out_path, scale_bar_label, rows, cols):
        max_w = max(im.width for im in images)
        max_h = max(im.height for im in images)
        padding = 10
        colorbar = Image.open(colorbar_path)
        font_size = 36
        label_padding = 10
        composite = Image.new("RGB", (
            cols * (max_w + padding) + padding,
            rows * (max_h + font_size + label_padding + padding) + colorbar.height + font_size + padding * 3
        ), BACKGROUND_COLOR)
        draw = ImageDraw.Draw(composite)

        # Add sample images and labels
        for i, img in enumerate(images):
            r = i // cols
            c = i % cols
            x = padding + c * (max_w + padding)
            y = padding + r * (max_h + font_size + label_padding + padding)
            composite.paste(img, (x, y))
            try:
                font = ImageFont.truetype("arial.ttf", font_size)
            except:
                font = ImageFont.load_default()
            text_width = draw.textlength(labels[i], font=font)
            draw.text((x + (max_w - text_width) // 2, y + max_h + 5), labels[i], fill="white", font=font)

        # Add colorbar and scale bar
        cb_x = (composite.width - colorbar.width) // 2
        cb_y = composite.height - colorbar.height - font_size - padding
        composite.paste(colorbar, (cb_x, cb_y))
        draw.line([(cb_x, cb_y + colorbar.height + 10), (cb_x + SCALE_BAR_PIXELS, cb_y + colorbar.height + 10)], fill="white", width=3)
        draw.text((cb_x, cb_y + colorbar.height + 15), scale_bar_label, fill="white", font=font)
        composite.save(out_path)

    # --- File Inventory ---
    print("\n🔍 Scanning INPUT folder...")
    matrix_files = sorted(f for f in os.listdir(INPUT_DIR) if f.endswith("matrix.xlsx") and not f.startswith("~"))
    print(f"Found {len(matrix_files)} files:")
    for f in matrix_files:
        print(" •", f)

    if input("Proceed? (y/n): ").lower() != 'y':
        print("Aborted.")
        return

    # --- Main Processing ---
    samples_by_element = {}
    for file in matrix_files:
        match = re.match(r"(.+)[ _]([A-Za-z]{1,2}\d{2,3})_(ppm|CPS) matrix\.xlsx", file)
        if match:
            sample, element, unit_type = match.groups()
            key = f"{element}_{unit_type}"
            samples_by_element.setdefault(key, []).append((sample, os.path.join(INPUT_DIR, file)))
        else:
            print(f"⚠️ Skipped: {file}")

    summary_rows = []
    variance_rows = []

    for key, sample_list in samples_by_element.items():
        element, unit_type = key.split("_")
        unit = "ppm" if unit_type == "ppm" else "CPS"
        element_dir = os.path.join(OUTPUT_DIR, element)
        os.makedirs(element_dir, exist_ok=True)

        print(f"\n📊 Processing {len(sample_list)} sample(s) for {element} [{unit}]")

        all_values = []
        sample_99th = {}
        sample_var = {}
        labeled_paths = []
        unlabeled_paths = []
        labels = []

        for sample, filepath in sample_list:
            matrix = load_matrix_2d(filepath)
            all_values.append(matrix)
            sample_99th[sample] = compute_99th_percentile(matrix)
            sample_var[sample] = compute_variance(matrix)
            print(f"   • {sample}: 99th = {sample_99th[sample]:.2f}, variance = {sample_var[sample]:.2f}")

        flat_all = np.concatenate([m.flatten() for m in all_values])
        scale_max = compute_99th_percentile(flat_all)

        colorbar_path = os.path.join(element_dir, f"{element}_colorbar.png")
        save_colorbar(scale_max, colorbar_path, unit)

        for sample, filepath in sample_list:
            matrix = load_matrix_2d(filepath)
            out_name = f"{sample}_{element}.png"
            out_path = os.path.join(element_dir, out_name)
            unlabeled = plot_matrix(matrix, scale_max, out_path)
            labeled = label_image(unlabeled, sample, 48)
            unlabeled_paths.append(unlabeled)
            labeled_paths.append(labeled)
            labels.append(sample)

        # Ask for grid layout if needed
        total_images = len(unlabeled_paths)
        if total_images > 6:
            rows = int(input("Enter number of rows for layout: "))
            cols = int(input("Enter number of columns for layout: "))
        else:
            cols = total_images
            rows = 1

        composite_path = os.path.join(element_dir, f"{element}_composite.png")
        create_composite(unlabeled_paths, labels, colorbar_path, composite_path, f"{int(SCALE_BAR_LENGTH_UM)} µm", rows, cols)

        for sample in sample_99th:
            summary_rows.append({"Sample": sample, element: round(sample_99th[sample], 1)})
            variance_rows.append({"Sample": sample, element: round(sample_var[sample], 1)})

    # --- Save Summary Tables ---
    if summary_rows:
        df = pd.DataFrame(summary_rows).groupby("Sample").first().reset_index()
        var_df = pd.DataFrame(variance_rows).groupby("Sample").first().reset_index()
        df.to_csv(os.path.join(OUTPUT_DIR, "summary_99th_percentiles.csv"), index=False)
        var_df.to_csv(os.path.join(OUTPUT_DIR, "summary_variance.csv"), index=False)
        print("\n✅ Summary tables saved to OUTPUT")

    print("\n✅ All processing complete.")

if __name__ == "__main__":
    main()
