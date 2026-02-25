# This script sets up a GUI application version of your command-line tool using tkinter
# Core features: Load data, choose element, layout options, preview, and save composite
# TP: THIS VERSION CLONED from Josh's most recent upload, with user-friendly GUI layout changes, 3 sig fig constraint, ppm or CPS input

import os
import sys
import tkinter as tk
from tkinter import filedialog, ttk, font, simpledialog
from . import custom_dialogs
import numpy as np  # pyright: ignore[reportMissingImports]
import matplotlib  # pyright: ignore[reportMissingImports]
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.colors import Normalize, LogNorm
from matplotlib import cm
from openpyxl import load_workbook
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
import math
import glob
import re
import tempfile
import time
from PIL import Image, ImageTk, ImageDraw, ImageFont
import shutil
import pandas as pd
import base64
import io

class CompositeApp:
    
    def _pt_from_font(self, var, default=16):
        """Return point size from a font StringVar: (None) -> default, else int in 6–72."""
        val = str(var.get()).strip()
        if not val or val == "(None)":
            return default
        try:
            return max(6, min(72, int(val)))
        except (ValueError, TypeError):
            return default

    def _pt_from_font_str(self, val_str, default=16):
        """Return point size from a string (e.g. from combobox.get()): (None) -> default, else int in 6–72."""
        val = str(val_str).strip()
        if not val or val == "(None)":
            return default
        try:
            return max(6, min(72, int(val)))
        except (ValueError, TypeError):
            return default

    def _get_element_label_font_value(self):
        """Return current Element label dropdown value from the widget so it's up-to-date when user just selected then clicked Preview."""
        c = getattr(self, "element_label_font_combobox", None)
        if c is not None and c.winfo_exists():
            return str(c.get()).strip()
        return str(self.element_label_font.get()).strip()

    def _on_element_label_font_change(self):
        """When Element label dropdown changes, update preview to show or hide label (if preview exists)."""
        if not getattr(self, "preview_image", None) or not getattr(self, "original_preview_image", None):
            return
        self.add_element_label()

    def _on_font_change_refresh_preview(self):
        """When any font dropdown changes and a preview exists, regenerate the composite so the new font sizes apply."""
        if not getattr(self, "matrices", None) or not self.matrices:
            return
        if not getattr(self, "preview_image", None):
            return
        # Regenerate preview with current font settings (sample names, scale bar, color bar, element label)
        self.generate_composite(preview=True)

    def get_contrasting_text_color(self, cmap_name):
        rgba = plt.get_cmap(cmap_name)(0.0)  # Color for 0 value
        r, g, b = rgba[:3]
        brightness = 0.299 * r + 0.587 * g + 0.114 * b
        return 'black' if brightness > 0.5 else 'white'
    
    def pseudolog_norm(self, vmin=1, vmax=100):
        """
        Returns a Normalize-like object for pseudolog scaling.
        This is a simple implementation: log(x+1) scaling, so that 0 maps to 0.
        """
        class PseudoLogNorm(Normalize):
            def __init__(self, vmin=None, vmax=None, clip=False):
                super().__init__(vmin, vmax, clip)
            def __call__(self, value, clip=None):
                vmin, vmax = self.vmin, self.vmax
                value = np.asarray(value)
                # Avoid negative/NaN
                value = np.where(value < 0, 0, value)
                # log1p for pseudo-log
                normed = (np.log1p(value) - np.log1p(vmin)) / (np.log1p(vmax) - np.log1p(vmin))
                return np.clip(normed, 0, 1)
            def inverse(self, value):
                vmin, vmax = self.vmin, self.vmax
                return np.expm1(value * (np.log1p(vmax) - np.log1p(vmin)) + np.log1p(vmin))
        return PseudoLogNorm(vmin=vmin, vmax=vmax)
    
    def parse_matrix_filename(self, filename):
        """
        Parse matrix filename to extract sample name, element, and unit type.
        Supports:
        1. Old format: {sample}[ _]{element}_{ppm|CPS} matrix.xlsx
           Element: 1-2 letters + 1-3 digits (e.g. Mo98, Ca44) OR summed channel TotalXxx (e.g. TotalMo).
        2. New format (Iolite raw CPS): {sample} {element} matrix.xlsx
        
        Returns: (sample, element, unit_type) or None if no match
        unit_type will be 'ppm', 'CPS', or 'raw' (for new format)
        """
        basename = os.path.basename(filename)
        # Element: standard [A-Za-z]{1,2}\d{1,3} or summed Total[A-Za-z]+ (e.g. TotalMo_ppm)
        elem_pattern = r"[A-Za-z]{1,2}\d{1,3}|Total[A-Za-z]+"
        
        # Try old format first: {sample}[ _]{element}_{ppm|CPS} matrix.xlsx
        match = re.match(rf"(.+?)[ _]({elem_pattern})_(ppm|CPS) matrix\.xlsx", basename)
        if match:
            sample, element, unit_type = match.groups()
            return (sample, element, unit_type)
        
        # Try new format: {sample} {element} matrix.xlsx
        match = re.match(rf"(.+?) ({elem_pattern}) matrix\.xlsx", basename)
        if match:
            sample, element = match.groups()
            return (sample, element, 'raw')
        
        return None

    def __init__(self, master):
        self.master = master
        master.title("ScaleBarOn Multi Map Scaler: v0.8.8")
        self._set_app_icon()

        self.pixel_size = tk.DoubleVar(value=6)
        self.scale_bar_length_um = tk.DoubleVar(value=500)
        self.num_rows = tk.IntVar(value=5)
        self.use_best_layout = tk.BooleanVar(value=False)  # Auto rows to minimize empty cells
        self.use_log = tk.BooleanVar(value=False)
        self.color_scheme = tk.StringVar(value="jet")
        self.element = tk.StringVar()
        self.unit = tk.StringVar()  # ppm, CPS, or raw; filters files when element has multiple units
        # Fonts: same style for each — (None) = off/default, then point sizes
        self.sample_name_font = tk.StringVar(value="12")   # (None) = off, else pt
        self.element_label_font = tk.StringVar(value="16")  # (None) = default 16 when adding, else pt
        self.scale_bar_font = tk.StringVar(value="10")     # (None) = off, else pt
        self.color_bar_font = tk.StringVar(value="10")     # (None) = default 10, else pt
        self.scale_max = tk.DoubleVar(value=1.0)  # New variable for scale_max, constrained to 2 decimal places
        self.use_custom_pixel_sizes = tk.BooleanVar(value=False)  # New variable for custom pixel sizes
        self.use_custom_pixel_sizes.trace_add('write', lambda *a: self._update_save_matrix_button_state())
        self.use_button_icons = tk.BooleanVar(value=False)  # Toggle for icon buttons
        # Optional credit text overlay on exported images (opt-in)
        self.add_credit_to_exports = tk.BooleanVar(value=False)
        self.credit_text = tk.StringVar(
            value=(
                "Collected at the Biomedical National Elemental Imaging Resource (BNEIR): "
                "NIGMS R24GM141194 & NIH 1S10OD032352"
            )
        )

        self.input_dir = None
        self.output_dir = "./OUTPUT"
        os.makedirs(self.output_dir, exist_ok=True)

        self.matrices = []
        self.labels = []
        self.preview_file = None
        self.preview_image = None  # Current preview image (may have labels)
        self.original_preview_image = None  # Original unlabeled preview image
        self.custom_pixel_sizes = {}  # Dictionary to store custom pixel sizes
        self.pixel_sizes_by_sample = {}
        
        # Progress tracking
        self.progress_samples = []  # List of sample names
        self.progress_elements = []  # List of element names (unique, for dropdown)
        self.progress_columns = []   # List of (element, unit_type) for table columns, grouped: ppm, then CPS, then raw
        self.progress_table = None
        self.progress_data = {}  # {(sample, element, unit_type): status} where status is 'complete', 'partial', 'missing_file', or None
        self.sample_include = {}  # sample name -> bool; which samples to include in scaling/preview (default True)
        self.sample_aliases = {}  # sample name -> alias used for labels/overlays (optional)
        
        # Status tracking for simplified log
        self.status = "Idle"  # Idle, Busy, Finishing
        
        # Button icons storage
        self.button_icons = {}  # Dictionary to store loaded button icons
        # BNEIR logo (GUI header)
        self.bneir_logo_image = None
        
        style = ttk.Style()
        style.configure("Hint.TLabel", foreground="gray", font=("TkDefaultFont", 12, "italic"))
        # Single GUI background grey so control panels and logo area match (no light grey box around logo)
        self._gui_bg = "#f0f0f0"
        style.configure("TFrame", background=self._gui_bg)
        # Load icons BEFORE creating buttons so they're available when buttons are created
        self.load_button_icons()
        # Load BNEIR logo if available
        self._load_bneir_logo()
        self.setup_widgets()

        # Load any existing sample aliases for this output directory
        self._load_sample_aliases()

    def setup_widgets(self):
        # Create notebook for tabs (similar to Muad'Data style)
        self.tabs = ttk.Notebook(self.master)
        self.tabs.pack(fill=tk.BOTH, expand=True)
        
        # Create tab frames
        self.setup_tab = ttk.Frame(self.tabs)
        self.preview_tab = ttk.Frame(self.tabs)
        
        # Add tabs
        self.tabs.add(self.setup_tab, text="Setup & Statistics")
        self.tabs.add(self.preview_tab, text="Preview & Export")
        
        # Build each tab
        self.build_setup_tab()
        self.build_preview_tab()
        
        # Set minimum window size
        self.master.minsize(600, 500)
        self.master.resizable(True, True)
        
        # Set initial window size to ensure all buttons and status log are visible
        # Width: 280px (left panel) + ~800px (right panel) + padding = ~1200px
        # Height: enough to show all controls and status log = ~800px
        self.master.geometry("1200x800")
        
        # Progress table window (created on demand - legacy, not used in tabbed interface)
        self.progress_table_window = None
        # Note: self.progress_table is created in build_setup_tab(), don't overwrite it here!
        
        # Preview window (created on demand) - kept for separate window option
        self.preview_window = None
        self.preview_window_label = None
    
    def build_setup_tab(self):
        """Build the Setup & Statistics tab."""
        # Left side: Controls (fixed width to match Preview tab)
        left_frame = ttk.Frame(self.setup_tab, width=280)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        left_frame.pack_propagate(False)  # Maintain fixed width
        # BNEIR logo at top of control panel (outside shaded groups)
        self._add_bneir_logo(left_frame)
        
        # Right side: Two-panel layout (Statistics Table, Progress Table)
        right_frame = ttk.Frame(self.setup_tab)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Select folders
        folders_group = ttk.LabelFrame(left_frame, text="Select folders", padding=10)
        folders_group.pack(fill=tk.X, pady=5)
        
        # Input folder
        ttk.Button(folders_group, text="Input", command=self.select_input_folder).pack(pady=(0, 2))
        self.input_folder_label = ttk.Label(folders_group, text="Input: None", font=("TkDefaultFont", 9), foreground="gray")
        self.input_folder_label.pack(pady=(0, 6))
        
        # Output folder
        ttk.Button(folders_group, text="Output", command=self.select_output_folder).pack(pady=(0, 2))
        self.output_folder_label = ttk.Label(folders_group, text=f"Output: {self.output_dir}", font=("TkDefaultFont", 9), foreground="gray")
        self.output_folder_label.pack(pady=(0, 0))
        
        # Element and Pixel Size controls
        element_pixel_frame = ttk.LabelFrame(left_frame, text="Element & Pixel Size", padding=10)
        element_pixel_frame.pack(fill=tk.X, pady=5)
        element_pixel_frame.columnconfigure(0, weight=1)
        element_pixel_frame.columnconfigure(1, weight=0, minsize=80)
        
        # Element dropdown
        ttk.Label(element_pixel_frame, text="Element:").grid(row=0, column=0, sticky="e", padx=5, pady=2)
        self.element_dropdown = ttk.Combobox(element_pixel_frame, textvariable=self.element, state="disabled", width=12)
        self.element_dropdown.grid(row=0, column=1, padx=5, pady=2, sticky="w")
        # Unit dropdown (ppm, CPS, raw); filters files when element has multiple units
        ttk.Label(element_pixel_frame, text="Unit:").grid(row=1, column=0, sticky="e", padx=5, pady=2)
        self.unit_dropdown = ttk.Combobox(element_pixel_frame, textvariable=self.unit, state="readonly", width=12)
        self.unit_dropdown.grid(row=1, column=1, padx=5, pady=2, sticky="w")
        # Update unit dropdown and statistics when element changes; statistics when unit changes
        self.element.trace('w', lambda *args: (self.update_unit_dropdown(), self.update_statistics_table()))
        self.unit.trace('w', lambda *args: self.update_statistics_table())
        
        # Pixel Size input
        ttk.Label(element_pixel_frame, text="Pixel Size (µm):").grid(row=2, column=0, sticky="e", padx=5, pady=2)
        pixel_entry_frame = ttk.Frame(element_pixel_frame)
        pixel_entry_frame.grid(row=2, column=1, sticky="w", padx=5, pady=2)
        ttk.Entry(pixel_entry_frame, textvariable=self.pixel_size, width=10).pack(side="top", anchor="w")
        ttk.Label(pixel_entry_frame, text="Hint: In your metadata", style="Hint.TLabel").pack(side="top", anchor="w", pady=(2, 0))
        
        # Multiple sizes checkbox
        multi_size_frame = ttk.Frame(element_pixel_frame)
        multi_size_frame.grid(row=3, column=0, columnspan=2, sticky="w", padx=5, pady=(5, 0))
        ttk.Checkbutton(multi_size_frame, text="Multiple sizes?", variable=self.use_custom_pixel_sizes).pack(side="top", anchor="w")
        ttk.Button(multi_size_frame, text="Pixel Sizes", command=self.handle_pixel_sizes).pack(side="top", anchor="w", pady=(5, 0))
        
        # Action buttons
        button_frame = ttk.LabelFrame(left_frame, text="Actions", padding=10)
        button_frame.pack(fill=tk.X, pady=5)
        
        # Calculate statistics button
        ttk.Label(button_frame, text="Calculate statistics", style="Hint.TLabel").pack(pady=(0, 2))
        summarize_icon = self.button_icons.get('summarize')
        if summarize_icon:
            self.summarize_btn = tk.Button(button_frame, image=summarize_icon, command=self.load_data, 
                                          padx=2, pady=8, bg='#f0f0f0', relief='raised',
                                          activebackground='#4CAF50')
            self.summarize_btn.image = summarize_icon
        else:
            self.summarize_btn = ttk.Button(button_frame, text="📊", command=self.load_data, width=1)
        self.summarize_btn.pack(pady=(0, 10))
        
        # Batch processing
        ttk.Label(button_frame, text="Batch Processing", style="Hint.TLabel").pack(pady=(0, 2))
        batch_icon = self.button_icons.get('batch')
        if batch_icon:
            self.batch_btn = tk.Button(button_frame, image=batch_icon,
                  command=self.batch_process_all_elements,
                  padx=2, pady=8, bg="#4CAF50", fg="black", relief='raised',
                  activebackground='#45a049')
            self.batch_btn.image = batch_icon
        else:
            self.batch_btn = tk.Button(button_frame, text="⚡", command=self.batch_process_all_elements, 
                  bg="#4CAF50", fg="black", width=1, height=3)
        self.batch_btn.pack(pady=(0, 5))
        
        # Shared progress bar (Batch processing and Calculate Statistics; updates in real time with Status Log)
        self.batch_progress_frame = ttk.Frame(button_frame)
        self.batch_progress_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(self.batch_progress_frame, text="Progress:", style="Hint.TLabel").pack(anchor="w")
        self.batch_progress_bar = ttk.Progressbar(self.batch_progress_frame, mode='determinate', maximum=100, value=0)
        self.batch_progress_bar.pack(fill=tk.X, pady=(2, 2))
        self.batch_progress_label = ttk.Label(self.batch_progress_frame, text="", style="Hint.TLabel")
        self.batch_progress_label.pack(anchor="w")
        
        # Progress Log at bottom of control panel
        log_group = ttk.LabelFrame(left_frame, text="Status Log", padding=5)
        log_group.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=False, padx=5, pady=5)
        
        log_frame = ttk.Frame(log_group)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        custom_font = font.Font(family="Arial", size=14, slant="roman")
        self.log = tk.Text(log_frame, height=12, width=30, wrap="word", font=custom_font)
        log_scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log.yview)
        self.log.configure(yscrollcommand=log_scrollbar.set)
        
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Use PanedWindow to create resizable sections
        paned = ttk.PanedWindow(right_frame, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Section 1: Statistics Table (top)
        stats_section = ttk.Frame(paned)
        paned.add(stats_section, weight=1)
        ttk.Label(stats_section, text="Statistics Table (Current Element):", style="Hint.TLabel").pack(anchor="center", padx=5, pady=(5, 0))
        
        stats_frame = ttk.Frame(stats_section)
        stats_frame.pack(fill=tk.BOTH, expand=True, pady=5, padx=5)
        
        # Create Treeview for statistics
        stats_columns = ('Sample', '25th Percentile', '50th Percentile', '75th Percentile', '99th Percentile', 'IQR', 'Mean')
        self.stats_table = ttk.Treeview(stats_frame, columns=stats_columns, show='headings', height=6)
        stats_xscroll = ttk.Scrollbar(stats_frame, orient=tk.HORIZONTAL, command=self.stats_table.xview)
        stats_yscroll = ttk.Scrollbar(stats_frame, orient=tk.VERTICAL, command=self.stats_table.yview)
        self.stats_table.configure(xscrollcommand=stats_xscroll.set, yscrollcommand=stats_yscroll.set)
        
        # Configure column headings
        for col in stats_columns:
            self.stats_table.heading(col, text=col)
            self.stats_table.column(col, width=100, anchor='center')
        
        self.stats_table.grid(row=0, column=0, sticky='nsew')
        stats_yscroll.grid(row=0, column=1, sticky='ns')
        stats_xscroll.grid(row=1, column=0, sticky='ew')
        stats_frame.grid_rowconfigure(0, weight=1)
        stats_frame.grid_columnconfigure(0, weight=1)
        
        # Section 2: Progress Table (bottom)
        progress_section = ttk.Frame(paned)
        paned.add(progress_section, weight=1)
        ttk.Label(progress_section, text="Progress Table:", style="Hint.TLabel").pack(anchor="center", padx=5, pady=(5, 0))

        progress_frame = ttk.Frame(progress_section)
        progress_frame.pack(fill=tk.BOTH, expand=True, pady=5, padx=5)

        # Progress table: scrollable grid of Labels so each cell can have its own colour (complete/partial/not started/no file)
        self.progress_table_canvas = tk.Canvas(progress_frame, highlightthickness=0)
        progress_yscroll = ttk.Scrollbar(progress_frame, orient=tk.VERTICAL, command=self.progress_table_canvas.yview)
        progress_xscroll = ttk.Scrollbar(progress_frame, orient=tk.HORIZONTAL, command=self.progress_table_canvas.xview)
        self.progress_table_canvas.configure(yscrollcommand=progress_yscroll.set, xscrollcommand=progress_xscroll.set)
        self.progress_table_inner = tk.Frame(self.progress_table_canvas, bg="#f0f0f0")
        self.progress_table_inner_window = self.progress_table_canvas.create_window((0, 0), window=self.progress_table_inner, anchor="nw")
        self.progress_table_canvas.grid(row=0, column=0, sticky="nsew")
        progress_yscroll.grid(row=0, column=1, sticky="ns")
        progress_xscroll.grid(row=1, column=0, sticky="ew")
        progress_frame.grid_rowconfigure(0, weight=1)
        progress_frame.grid_columnconfigure(0, weight=1)

        def _on_progress_frame_configure(event):
            self.progress_table_canvas.configure(scrollregion=self.progress_table_canvas.bbox("all"))
        def _on_canvas_configure(event):
            self.progress_table_canvas.itemconfig(self.progress_table_inner_window, width=event.width)
        self.progress_table_inner.bind("<Configure>", _on_progress_frame_configure)
        self.progress_table_canvas.bind("<Configure>", _on_canvas_configure)
        self.progress_table = self.progress_table_inner  # so "if progress_table" checks still work

        # Legend for progress table colours
        legend_frame = ttk.Frame(progress_section)
        legend_frame.pack(fill=tk.X, padx=5, pady=(2, 0))
        _legends = (
            ("#90EE90", "Complete (composite saved)"),
            ("#FFE4B5", "Partial (statistics only)"),
            ("#FFB6C1", "Not started"),
            ("#D3D3D3", "No input file (!)"),
        )
        for bg, text in _legends:
            row = ttk.Frame(legend_frame)
            row.pack(side=tk.LEFT, padx=(0, 12))
            swatch = tk.Frame(row, width=14, height=14, bg=bg, relief="solid", borderwidth=1)
            swatch.pack(side=tk.LEFT, padx=(0, 4))
            swatch.pack_propagate(False)
            ttk.Label(row, text=text, style="Hint.TLabel").pack(side=tk.LEFT)

        # Include column cells are bound to _toggle_sample_include in update_progress_table()

        # Add refresh button for progress table
        refresh_frame = ttk.Frame(progress_section)
        refresh_frame.pack(fill=tk.X, padx=5, pady=(0, 5))
        ttk.Button(refresh_frame, text="Refresh Progress Table", command=self.refresh_progress_table).pack(side=tk.RIGHT)
    
    def build_preview_tab(self):
        """Build the Preview & Export tab."""
        # Left side: Controls (fixed width to match Setup tab)
        control_frame = ttk.Frame(self.preview_tab, width=280)
        control_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        control_frame.pack_propagate(False)  # Maintain fixed width
        # BNEIR logo at top of control panel (outside shaded groups)
        self._add_bneir_logo(control_frame)
        
        # Right side: Preview pane
        preview_frame = ttk.Frame(self.preview_tab)
        preview_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Element and Unit dropdowns at top (stacked, with aligned comboboxes)
        element_group = ttk.Frame(control_frame)
        element_group.pack(pady=5, fill=tk.X)

        # Element row
        ttk.Label(element_group, text="Element:").grid(row=0, column=0, sticky="e", padx=5, pady=(0, 2))
        # Element strings are short (e.g. 'Ca44', 'Total Ca'), so a narrow width is sufficient
        self.element_dropdown_preview = ttk.Combobox(
            element_group, textvariable=self.element, state="readonly", width=10
        )
        self.element_dropdown_preview.grid(row=0, column=1, sticky="w", padx=5, pady=(0, 2))

        # Unit row (moved below element to prevent being cut off)
        ttk.Label(element_group, text="Unit:").grid(row=1, column=0, sticky="e", padx=5, pady=(0, 0))
        self.unit_dropdown_preview = ttk.Combobox(
            element_group, textvariable=self.unit, state="readonly", width=10
        )
        self.unit_dropdown_preview.grid(row=1, column=1, sticky="w", padx=5, pady=(0, 0))
        
        # Layout controls
        layout_frame = ttk.LabelFrame(control_frame, text="Layout", padding=10)
        layout_frame.pack(fill=tk.X, pady=5)
        layout_frame.columnconfigure(0, weight=1)
        layout_frame.columnconfigure(1, weight=0, minsize=80)
        
        # Scale bar length
        ttk.Label(layout_frame, text="Scale bar length (µm):").grid(row=0, column=0, sticky="e", padx=5, pady=2)
        scale_bar_entry = ttk.Entry(layout_frame, textvariable=self.scale_bar_length_um, width=8)
        scale_bar_entry.grid(row=0, column=1, padx=5, pady=2, sticky="w")
        
        # Rows
        ttk.Label(layout_frame, text="Rows:").grid(row=1, column=0, sticky="e", padx=5, pady=2)
        rows_entry = ttk.Entry(layout_frame, textvariable=self.num_rows, width=8)
        rows_entry.grid(row=1, column=1, padx=5, pady=2, sticky="w")
        
        # Use best layout (auto rows to minimize blank cells; same as batch)
        ttk.Checkbutton(layout_frame, text="Use best layout (auto rows)", variable=self.use_best_layout).grid(row=2, column=0, columnspan=2, sticky="w", padx=5, pady=2)
        
        # Display controls
        display_frame = ttk.LabelFrame(control_frame, text="Display", padding=10)
        display_frame.pack(fill=tk.X, pady=5)
        display_frame.columnconfigure(0, weight=1)
        display_frame.columnconfigure(1, weight=0, minsize=80)
        
        # Color scheme
        ttk.Label(display_frame, text="Color Scheme:").grid(row=0, column=0, sticky="e", padx=5, pady=2)
        self.color_scheme_dropdown = ttk.Combobox(display_frame, textvariable=self.color_scheme, values=plt.colormaps(), width=8)
        self.color_scheme_dropdown.grid(row=0, column=1, padx=5, pady=2, sticky="w")
        
        # Scale max
        ttk.Label(display_frame, text="Scale Max:").grid(row=1, column=0, sticky="e", padx=5, pady=2)
        scale_max_entry = ttk.Entry(display_frame, textvariable=self.scale_max, width=8)
        scale_max_entry.grid(row=1, column=1, padx=5, pady=2, sticky="w")
        
        # Log Scale
        ttk.Checkbutton(display_frame, text="Log Scale", variable=self.use_log).grid(row=2, column=0, columnspan=2, pady=2)
        
        # Label font size: same style for each — (None) = off/default, then point sizes
        font_frame = ttk.LabelFrame(control_frame, text="Label font size", padding=10)
        font_frame.pack(fill=tk.X, pady=5)
        font_frame.columnconfigure(0, weight=1)
        font_frame.columnconfigure(1, weight=0, minsize=80)
        font_dropdown_values = ["(None)", "8", "9", "10", "12", "14", "16", "18", "20", "24"]
        def _make_font_row(row, label, var, width=8):
            ttk.Label(font_frame, text=label).grid(row=row, column=0, sticky="e", padx=5, pady=2)
            c = ttk.Combobox(font_frame, textvariable=var, values=font_dropdown_values, width=width, state="readonly")
            c.grid(row=row, column=1, padx=5, pady=2, sticky="w")
            return c
        _make_font_row(0, "Sample names:", self.sample_name_font)
        self.element_label_font_combobox = _make_font_row(1, "Element label:", self.element_label_font)
        _make_font_row(2, "Scale bar:", self.scale_bar_font)
        _make_font_row(3, "Color bar:", self.color_bar_font)
        ttk.Checkbutton(font_frame, text="Add BNEIR credit", variable=self.add_credit_to_exports).grid(row=4, column=0, columnspan=2, sticky="", padx=5, pady=(4, 0))
        # When any font dropdown or credit checkbox changes, refresh preview if one exists
        def _refresh_preview_on_font_change(*args):
            self._on_font_change_refresh_preview()
        self.element_label_font.trace_add("write", _refresh_preview_on_font_change)
        self.sample_name_font.trace_add("write", _refresh_preview_on_font_change)
        self.scale_bar_font.trace_add("write", _refresh_preview_on_font_change)
        self.color_bar_font.trace_add("write", _refresh_preview_on_font_change)
        self.add_credit_to_exports.trace_add("write", _refresh_preview_on_font_change)

        # Action buttons
        action_frame = ttk.Frame(control_frame)
        action_frame.pack(fill=tk.X, pady=10)
        
        # Preview button
        ttk.Label(action_frame, text="Preview composite", style="Hint.TLabel").pack(pady=(5, 0))
        preview_icon = self.button_icons.get('preview')
        if preview_icon:
            self.preview_btn = tk.Button(action_frame, image=preview_icon, command=self.preview_composite,
                                        padx=2, pady=8, bg='#f0f0f0', relief='raised',
                                        activebackground='#4CAF50')
            self.preview_btn.image = preview_icon
        else:
            self.preview_btn = ttk.Button(action_frame, text="👁️", command=self.preview_composite, width=1)
        self.preview_btn.pack(pady=(0, 10))
        
        # Save Composite button
        ttk.Label(action_frame, text="Save Composite", style="Hint.TLabel").pack(pady=(5, 0))
        save_icon = self.button_icons.get('save')
        if save_icon:
            self.save_btn = tk.Button(action_frame, image=save_icon, command=self.save_composite,
                                     padx=2, pady=8, bg='#f0f0f0', relief='raised',
                                     activebackground='#4CAF50')
            self.save_btn.image = save_icon
        else:
            self.save_btn = ttk.Button(action_frame, text="💾", command=self.save_composite, width=1)
        self.save_btn.pack(pady=(0, 10))
        
        # Save Composite Matrix button (for Muad'Data); disabled when multiple pixel sizes (file has no per-panel scale)
        ttk.Label(action_frame, text="Save Composite Matrix (for Muad'Data)", style="Hint.TLabel").pack(pady=(5, 0))
        save_matrix_icon = self.button_icons.get('save_matrix')
        if save_matrix_icon:
            self.save_matrix_btn = tk.Button(action_frame, image=save_matrix_icon, command=self.save_composite_matrix,
                                             padx=2, pady=8, bg='#f0f0f0', relief='raised',
                                             activebackground='#4CAF50')
            self.save_matrix_btn.image = save_matrix_icon
        else:
            self.save_matrix_btn = ttk.Button(action_frame, text="💾 Matrix", command=self.save_composite_matrix, width=15)
        self.save_matrix_btn.pack(pady=(0, 5))
        self._update_save_matrix_button_state()
        
        # Progress Log at bottom of control panel
        log_group_preview = ttk.LabelFrame(control_frame, text="Status Log", padding=5)
        log_group_preview.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=False, padx=5, pady=5)
        
        log_frame_preview = ttk.Frame(log_group_preview)
        log_frame_preview.pack(fill=tk.BOTH, expand=True)
        
        custom_font_preview = font.Font(family="Arial", size=14, slant="roman")
        self.log_preview = tk.Text(log_frame_preview, height=12, width=30, wrap="word", font=custom_font_preview)
        log_scrollbar_preview = ttk.Scrollbar(log_frame_preview, orient=tk.VERTICAL, command=self.log_preview.yview)
        self.log_preview.configure(yscrollcommand=log_scrollbar_preview.set)
        
        self.log_preview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scrollbar_preview.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Preview pane
        self.preview_container = ttk.Frame(preview_frame)
        self.preview_container.pack(fill=tk.BOTH, expand=True)
        
        self.preview_label = ttk.Label(self.preview_container, 
                                      text="\n\nCalculate statistics in the 'Setup & Statistics' tab first,\nthen click 'Preview composite' to generate preview here.\n\n",
                                      font=("Arial", 12), foreground="gray", justify=tk.CENTER, anchor="center")
        self.preview_label.pack(fill=tk.BOTH, expand=True)
        
        # Bind resize event to update the preview image's aspect ratio
        self.preview_container.bind("<Configure>", self.on_resize)

    def _load_bneir_logo(self):
        """Load BNEIR logo image (used as small header in control panels) if file is available."""
        self.bneir_logo_image = None
        try:
            # Prefer package-relative icons folder; fall back to user-specific path if present
            candidates = [
                os.path.join(os.path.dirname(__file__), "icons", "BNEIR_logo.png"),
                os.path.join(os.path.dirname(__file__), "icons", "BNEIR_logo_1.png"),
                # Fallback to user workspace asset path (development)
                "/Users/d19766d/.cursor/projects/Users-d19766d-Documents-GitHub-scalebaron/assets/BNEIR_logo_1-ab04aad1-b6cf-48ac-a035-2a5fcce0d8c8.png",
            ]
            logo_path = None
            for path in candidates:
                if os.path.exists(path):
                    logo_path = path
                    break
            if not logo_path:
                return
            img = Image.open(logo_path)
            max_width = 220
            w, h = img.size
            if w > max_width:
                new_w = max_width
                new_h = int(h * new_w / float(w))
                img = img.resize((new_w, new_h), Image.LANCZOS)
            self.bneir_logo_image = ImageTk.PhotoImage(img)
        except Exception:
            self.bneir_logo_image = None

    def _add_bneir_logo(self, parent):
        """Add BNEIR logo at the top of a control panel if logo is loaded. Centered; uses same GUI default grey as panels."""
        if not getattr(self, "bneir_logo_image", None):
            return
        bg = getattr(self, "_gui_bg", "#f0f0f0")
        container = tk.Frame(parent, bg=bg)
        container.pack(pady=(0, 5))
        lbl = tk.Label(container, image=self.bneir_logo_image, bg=bg)
        lbl.pack(anchor=tk.CENTER)

    def _set_app_icon(self):
        """Set the ScaleBarOn logo for the main window and dialogs (replaces default Python logo)."""
        try:
            candidates = [
                os.path.join(os.path.dirname(__file__), "icons", "scalebaron_icon.png"),
                os.path.join(os.path.dirname(__file__), "icons", "ScaleBarOn LOGO.png"),
            ]
            icon_path = None
            for path in candidates:
                if os.path.exists(path):
                    icon_path = path
                    break
            if not icon_path:
                return
            img = Image.open(icon_path)
            if img.size[0] > 64 or img.size[1] > 64:
                img = img.resize((64, 64), Image.LANCZOS)
            self._app_icon = ImageTk.PhotoImage(img)
            self.master.iconphoto(True, self._app_icon)
            self.master._dialog_icon = self._app_icon  # For custom_dialogs to show logo inside dialogs
            self.master._app_icon_ref = self._app_icon  # Keep ref on root so icon persists
            # On macOS, iconphoto() only affects the window; set dock icon via Cocoa (requires PyObjC)
            if sys.platform == "darwin":
                try:
                    from AppKit import NSApplication, NSImage
                    nsimg = NSImage.alloc().initWithContentsOfFile_(icon_path)
                    if nsimg is not None:
                        NSApplication.sharedApplication().setApplicationIconImage_(nsimg)
                except Exception:
                    pass
        except Exception:
            pass

    def load_button_icons(self):
        """
        Load custom button icons from the icons folder if they exist.
        Falls back to embedded base64-encoded icons if external files are not found.
        """
        icons_dir = os.path.join(os.path.dirname(__file__), 'icons')
        icon_files = {
            'summarize': 'summarize.png',
            'preview': 'preview.png',
            'add_label': 'add_label.png',  # Also check for 'label.png' as fallback
            'save': 'save.png',
            'save_matrix': 'save_matrix.png',
            'batch': 'batch.png',
            'progress': 'progress.png'
        }
        
        # Try to load embedded icons as fallback
        try:
            # Try relative import first (when used as package)
            try:
                from . import embedded_icons
                embedded_icons_dict = embedded_icons.EMBEDDED_ICONS
            except ImportError:
                # Try absolute import (when run directly)
                import embedded_icons
                embedded_icons_dict = embedded_icons.EMBEDDED_ICONS
        except (ImportError, AttributeError):
            # If embedded_icons module doesn't exist, use empty dict
            embedded_icons_dict = {}
        
        def load_icon_from_data(icon_data, source_name):
            """Load icon from image data (either file path or base64 string)."""
            try:
                if isinstance(icon_data, str) and len(icon_data) > 100 and not os.path.exists(icon_data):
                    # Base64 string (from embedded_icons)
                    icon_bytes = base64.b64decode(icon_data)
                    icon_img = Image.open(io.BytesIO(icon_bytes))
                else:
                    # File path
                    icon_img = Image.open(icon_data)
                
                # Resize to 32x32 (larger icons for better visibility)
                if icon_img.size[0] > 32 or icon_img.size[1] > 32:
                    icon_img = icon_img.resize((32, 32), Image.LANCZOS)
                elif icon_img.size[0] < 32 or icon_img.size[1] < 32:
                    # Upscale smaller icons
                    icon_img = icon_img.resize((32, 32), Image.LANCZOS)
                
                # Convert to PhotoImage for Tkinter
                return ImageTk.PhotoImage(icon_img)
            except Exception as e:
                # Silently fail - will use Unicode fallback
                return None
        
        for key, filename in icon_files.items():
            icon_loaded = False
            
            # First, try to load from external file
            icon_path = os.path.join(icons_dir, filename)
            # Special case: also check for 'label.png' if 'add_label.png' not found
            if key == 'add_label' and not os.path.exists(icon_path):
                alt_path = os.path.join(icons_dir, 'label.png')
                if os.path.exists(alt_path):
                    icon_path = alt_path
            
            if os.path.exists(icon_path):
                photo_image = load_icon_from_data(icon_path, filename)
                if photo_image:
                    self.button_icons[key] = photo_image
                    icon_loaded = True
            
            # If external file not found or failed, try embedded icon
            if not icon_loaded and key in embedded_icons_dict:
                photo_image = load_icon_from_data(embedded_icons_dict[key], f"embedded {key}")
                if photo_image:
                    self.button_icons[key] = photo_image
                    icon_loaded = True
            
            # If icon not loaded, will silently use Unicode/text fallback

    def select_input_folder(self):
        folder_selected = filedialog.askdirectory(initialdir=self.input_dir or ".", title="Select Input Directory")
        if folder_selected:
            self.input_dir = folder_selected
            self.log_print(f"Input folder updated to: {self.input_dir}")
            self.update_element_dropdown()
            if hasattr(self, 'input_folder_label'):
                self.input_folder_label.config(text=f"Input: {self.input_dir}")
            # Scan and populate progress table
            self.scan_progress_table()

    def select_output_folder(self):
        folder_selected = filedialog.askdirectory(initialdir=self.output_dir, title="Select Output Directory")
        if folder_selected:
            self.output_dir = folder_selected
            self.output_folder_label.config(text=f"Output: {self.output_dir}")
            self.log_print(f"Output folder updated to: {self.output_dir}")
            # Reload sample aliases for the new output directory
            self.sample_aliases.clear()
            self._load_sample_aliases()
            # Re-check existing progress and update table
            if self.progress_samples:
                self._check_existing_progress()
                # Only update table if widget exists
                if hasattr(self, 'progress_table') and self.progress_table is not None:
                    self.update_progress_table()
        # If user cancels, keep the current output directory (no change)

    def show_progress_table_window(self):
        """Refresh the embedded progress table (now in Setup & Statistics tab)."""
        # Debug: Check if progress_table exists
        if not hasattr(self, 'progress_table') or self.progress_table is None:
            self.log_print("⚠️ Progress table widget not initialized. Please restart the application.")
            return
        
        self.set_status("Busy")
        
        # First scan the input folder to build the progress table structure
        if not self.progress_samples or not self.progress_elements:
            if self.input_dir and os.path.isdir(self.input_dir):
                self.log_print("Status: Busy - Scanning input folder...")
                self.scan_progress_table()
            elif self.output_dir and os.path.isdir(self.output_dir):
                # If no input dir but output dir exists, try to infer from output structure
                self.log_print("Status: Busy - Scanning output folder...")
                self._scan_progress_from_output()
            else:
                self.log_print("⚠️ Please select an input or output folder first to view progress table.")
                self.set_status("Idle")
                return
        
        # Always refresh the status to get latest updates
        self._check_existing_progress()
        self.update_progress_table()
        
        # Debug info
        if self.progress_samples and self.progress_elements:
            self.log_print(f"Status: Idle - Progress table updated: {len(self.progress_samples)} samples, {len(self.progress_elements)} elements")
        else:
            self.log_print("⚠️ No progress data found. Please calculate statistics first.")
        
        self.set_status("Idle")
        
        # Switch to Setup & Statistics tab to show the table
        self.tabs.select(0)  # Index 0 is the Setup & Statistics tab
    
    def refresh_progress_table(self):
        """Refresh the progress table by re-scanning and checking status."""
        if not hasattr(self, 'progress_table') or self.progress_table is None:
            self.log_print("⚠️ Progress table widget not initialized.")
            # Try to initialize if not already done
            if self.input_dir and os.path.isdir(self.input_dir):
                self.scan_progress_table()
            elif self.output_dir and os.path.isdir(self.output_dir):
                self._scan_progress_from_output()
            return
        
        self.set_status("Busy")
        self.log_print("Status: Busy - Refreshing progress table...")
        
        # Re-scan and update
        if self.input_dir and os.path.isdir(self.input_dir):
            self.scan_progress_table()
        elif self.output_dir and os.path.isdir(self.output_dir):
            # If we have progress data, just refresh status
            if self.progress_samples and self.progress_elements:
                self._check_existing_progress()
                self.update_progress_table()
            else:
                # Try to infer from output
                self._scan_progress_from_output()
        else:
            self.log_print("⚠️ No input or output folder set.")
            self.set_status("Idle")
            return
        
        self.set_status("Idle")
        if self.progress_samples and self.progress_elements:
            self.log_print(f"Status: Idle - Progress table refreshed: {len(self.progress_samples)} samples, {len(self.progress_elements)} elements")
        else:
            self.log_print("Status: Idle - No progress data found.")
    
    def _on_progress_table_click(self, event):
        """Toggle sample include when user clicks the Include (checkbox) column. No-op when using grid (Include cells bound separately)."""
        if not hasattr(self, 'progress_table') or self.progress_table is None:
            return
        if not hasattr(self.progress_table, 'identify_region'):
            return  # Grid of Labels, not Treeview; Include handled by _toggle_sample_include
        region = self.progress_table.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.progress_table.identify_column(event.x)
        # Include is first data column (#1)
        if col != "#1":
            return
        item = self.progress_table.identify_row(event.y)
        if not item:
            return
        try:
            values = self.progress_table.item(item, "values")
            if len(values) < 2:
                return
            sample = values[1]
            if sample not in self.progress_samples:
                return
            self.sample_include[sample] = not self.sample_include.get(sample, True)
            self.update_progress_table()
        except Exception:
            pass
    
    def get_selected_samples(self):
        """Return list of sample names that are selected for scaling/preview. If none selected, returns all (backward compat)."""
        selected = [s for s in self.progress_samples if self.sample_include.get(s, True)]
        if not selected:
            return list(self.progress_samples)
        return selected
    
    def scan_progress_table(self):
        """Scan input folder and build initial progress table. Columns grouped by unit: ppm, CPS, raw."""
        if not self.input_dir or not os.path.isdir(self.input_dir):
            return

        samples = set()
        columns_seen = set()  # (element, unit_type)
        files_found = set()   # (sample, element, unit_type)

        # Scan for all matrix files (old format ppm/CPS, new format raw)
        for file in glob.glob(os.path.join(self.input_dir, "* matrix.xlsx")):
            parsed = self.parse_matrix_filename(file)
            if parsed:
                sample, element, unit_type = parsed
                samples.add(sample)
                columns_seen.add((element, unit_type))
                files_found.add((sample, element, unit_type))

        def elem_sort_key(elem):
            m = re.search(r"(\D+)(\d+)$", elem)
            if m:
                return (m.group(1), int(m.group(2)))
            return (elem, 0)

        unit_order = ('ppm', 'CPS', 'raw')
        # Build progress_columns: ppm first, then CPS, then raw; within each group sort by element
        by_unit = {}
        for (element, unit_type) in columns_seen:
            by_unit.setdefault(unit_type, []).append(element)
        self.progress_columns = []
        for ut in unit_order:
            if ut in by_unit:
                for elem in sorted(by_unit[ut], key=elem_sort_key):
                    self.progress_columns.append((elem, ut))
        # Any unit type not in unit_order (e.g. future) at end
        for ut, elems in by_unit.items():
            if ut not in unit_order:
                for elem in sorted(elems, key=elem_sort_key):
                    self.progress_columns.append((elem, ut))

        self.progress_samples = sorted(samples)
        self.progress_elements = sorted(set(e for e, _ in self.progress_columns), key=elem_sort_key)
        for sample in self.progress_samples:
            if sample not in self.sample_include:
                self.sample_include[sample] = True

        # Build progress_data: (sample, element, unit_type) -> status
        self.progress_data = {}
        for sample in self.progress_samples:
            for (element, unit_type) in self.progress_columns:
                if (sample, element, unit_type) in files_found:
                    self.progress_data[(sample, element, unit_type)] = None
                else:
                    self.progress_data[(sample, element, unit_type)] = 'missing_file'

        self._check_existing_progress()
        # Only update table if widget exists
        if hasattr(self, 'progress_table') and self.progress_table is not None:
            self.update_progress_table()
    
    def _scan_progress_from_output(self):
        """Scan output folder to infer progress when input folder is not available. Folders are element_unit (e.g. Fe_ppm)."""
        if not self.output_dir or not os.path.isdir(self.output_dir):
            return
        
        samples = set()
        progress_cols = []  # (element, unit_type)
        
        # Scan output directories (element_unit folders e.g. Fe_ppm, Fe_CPS)
        for item in os.listdir(self.output_dir):
            element_dir = os.path.join(self.output_dir, item)
            if os.path.isdir(element_dir):
                composite_path = os.path.join(element_dir, f"{item}_composite.png")
                hist_dir = os.path.join(element_dir, 'Histograms')
                
                if os.path.exists(composite_path) or (os.path.isdir(hist_dir) and os.listdir(hist_dir)):
                    # Parse element_unit: "Fe_ppm" -> (Fe, ppm)
                    if "_" in item and item.rsplit("_", 1)[1] in ('ppm', 'CPS', 'raw'):
                        elem, ut = item.rsplit("_", 1)
                        progress_cols.append((elem, ut))
                    else:
                        progress_cols.append((item, 'ppm'))  # legacy: no unit in folder name
                    if os.path.isdir(hist_dir):
                        for hist_file in os.listdir(hist_dir):
                            if hist_file.endswith('_histogram.png'):
                                sample = hist_file.replace('_histogram.png', '')
                                samples.add(sample)
        
        if samples and progress_cols:
            def sort_key(elem):
                m = re.search(r"(\D+)(\d+)$", elem)
                if m:
                    return (m.group(1), int(m.group(2)))
                return (elem, 0)

            self.progress_samples = sorted(samples)
            self.progress_columns = sorted(set(progress_cols), key=lambda x: (('ppm', 'CPS', 'raw').index(x[1]) if x[1] in ('ppm', 'CPS', 'raw') else 99, sort_key(x[0])))
            self.progress_elements = sorted(set(e for e, _ in self.progress_columns), key=sort_key)
            for sample in self.progress_samples:
                if sample not in self.sample_include:
                    self.sample_include[sample] = True
            self.progress_data = {}
            for sample in self.progress_samples:
                for (element, unit_type) in self.progress_columns:
                    self.progress_data[(sample, element, unit_type)] = None

            self._check_existing_progress()
            if hasattr(self, 'progress_table') and self.progress_table is not None:
                self.update_progress_table()
            self.log_print(f"📊 Inferred progress from output: {len(self.progress_samples)} samples, {len(self.progress_elements)} elements")
    
    def _check_existing_progress(self):
        """Check output folder for existing files. Supports both new (element_unit) and legacy (element) folder structure."""
        if not self.output_dir or not os.path.isdir(self.output_dir):
            return

        elem_units_with_composite = set()
        for (element, unit_type) in self.progress_columns:
            # New format: output_dir/Fe_ppm/Fe_ppm_composite.png
            elem_out = f"{element}_{unit_type}"
            new_dir = os.path.join(self.output_dir, elem_out)
            new_composite = os.path.join(new_dir, f"{elem_out}_composite.png")
            # Legacy format: output_dir/Fe/Fe_composite.png
            legacy_dir = os.path.join(self.output_dir, element)
            legacy_composite = os.path.join(legacy_dir, f"{element}_composite.png")
            if (os.path.exists(new_composite) and os.path.getsize(new_composite) > 0) or \
               (os.path.exists(legacy_composite) and os.path.getsize(legacy_composite) > 0):
                elem_units_with_composite.add((element, unit_type))

        for (sample, element, unit_type), current_status in list(self.progress_data.items()):
            if current_status == 'missing_file':
                continue
            if (element, unit_type) in elem_units_with_composite:
                self.progress_data[(sample, element, unit_type)] = 'complete'
            else:
                # Check for partial (histogram exists): new or legacy path
                elem_out = f"{element}_{unit_type}"
                new_hist = os.path.join(self.output_dir, elem_out, 'Histograms', f"{sample}_histogram.png")
                legacy_hist = os.path.join(self.output_dir, element, 'Histograms', f"{sample}_histogram.png")
                if (os.path.exists(new_hist) and os.path.getsize(new_hist) > 0) or \
                   (os.path.exists(legacy_hist) and os.path.getsize(legacy_hist) > 0):
                    self.progress_data[(sample, element, unit_type)] = 'partial'
                else:
                    self.progress_data[(sample, element, unit_type)] = None
    
    def update_statistics_table(self, stats_df=None):
        """Update the statistics table display with current element's statistics."""
        if not hasattr(self, 'stats_table') or self.stats_table is None:
            return
        
        # Clear existing rows
        for item in self.stats_table.get_children():
            self.stats_table.delete(item)
        
        # If no stats_df provided, try to load from file
        if stats_df is None:
            elem_out = self.get_element_output_subdir()
            if not elem_out:
                return
            
            stats_path = os.path.join(self.output_dir, elem_out, f"{elem_out}_statistics.csv")
            if os.path.exists(stats_path):
                try:
                    stats_df = pd.read_csv(stats_path)
                    # Ingest any stored aliases from the statistics file
                    if 'Sample' in stats_df.columns and 'Alias' in stats_df.columns:
                        for _, row in stats_df[['Sample', 'Alias']].dropna().iterrows():
                            sample = str(row['Sample'])
                            alias = str(row['Alias']).strip()
                            if alias and alias != sample:
                                self.sample_aliases[sample] = alias
                except Exception as e:
                    self.log_print(f"⚠️ Could not load statistics: {e}")
                    return
            else:
                return
        
        # Populate table
        for _, row in stats_df.iterrows():
            values = (
                str(row.get('Sample', '')),
                f"{row.get('25th Percentile', 0):.2f}",
                f"{row.get('50th Percentile', 0):.2f}",
                f"{row.get('75th Percentile', 0):.2f}",
                f"{row.get('99th Percentile', 0):.2f}",
                f"{row.get('IQR', 0):.2f}",
                f"{row.get('Mean', 0):.2f}"
            )
            self.stats_table.insert('', tk.END, values=values)
    
    def _toggle_sample_include(self, sample):
        """Toggle include state for a sample (called when user clicks Include cell)."""
        if sample not in self.progress_samples:
            return
        self.sample_include[sample] = not self.sample_include.get(sample, True)
        self.update_progress_table()

    def update_progress_table(self):
        """Update the progress table display. Uses a grid of Labels so each cell has its own colour."""
        if not hasattr(self, 'progress_table_inner') or self.progress_table_inner is None:
            return

        inner = self.progress_table_inner
        for w in inner.winfo_children():
            w.destroy()

        # Status colours (per-cell)
        _bg_complete = "#90EE90"
        _bg_partial = "#FFE4B5"
        _bg_missing = "#FFB6C1"
        _bg_missing_file = "#D3D3D3"
        _bg_header = "#e0e0e0"
        _bg_include_sample = "#f0f0f0"

        progress_cols = getattr(self, 'progress_columns', None) or [(e, '') for e in self.progress_elements]
        if not progress_cols:
            tk.Label(inner, text="No elements found. Please calculate statistics first.", bg=_bg_header, font=("TkDefaultFont", 11)).grid(row=0, column=0, sticky="ew", padx=2, pady=2)
            inner.update_idletasks()
            if hasattr(self, 'progress_table_canvas'):
                self.progress_table_canvas.configure(scrollregion=self.progress_table_canvas.bbox("all"))
            return

        if not self.progress_samples:
            tk.Label(inner, text="No samples found. Please calculate statistics first.", bg=_bg_header, font=("TkDefaultFont", 11)).grid(row=0, column=0, sticky="ew", padx=2, pady=2)
            inner.update_idletasks()
            if hasattr(self, 'progress_table_canvas'):
                self.progress_table_canvas.configure(scrollregion=self.progress_table_canvas.bbox("all"))
            return
        max_sample_len = max([len(s) for s in self.progress_samples] + [6])
        max_alias_len = max([len(self.sample_aliases.get(s, "")) for s in self.progress_samples] + [5])
        col_widths = [
            4,  # Include checkbox
            max(8, min(max_sample_len + 1, 24)),   # Sample ID
            max(8, min(max_alias_len + 1, 24)),    # Alias
        ] + [5] * len(progress_cols)

        # Header row 0: unit labels grouped over element columns (PPM | CPS | RAW)
        tk.Label(inner, text="", bg=_bg_header, font=("TkDefaultFont", 11, "bold"), width=col_widths[0], anchor="center").grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        tk.Label(inner, text="", bg=_bg_header, font=("TkDefaultFont", 11, "bold"), width=col_widths[1], anchor="w").grid(row=0, column=1, sticky="nsew", padx=1, pady=1)
        tk.Label(inner, text="", bg=_bg_header, font=("TkDefaultFont", 11, "bold"), width=col_widths[2], anchor="w").grid(row=0, column=2, sticky="nsew", padx=1, pady=1)
        col_start = 3
        idx = 0
        while idx < len(progress_cols):
            element, unit_type = progress_cols[idx]
            group_ut = unit_type or ""
            run_start = idx
            while idx < len(progress_cols) and progress_cols[idx][1] == unit_type:
                idx += 1
            span = idx - run_start
            label_text = group_ut.upper() if group_ut else ""
            if label_text:
                # Compute approximate total width of this unit group for nicer centering
                group_width = sum(col_widths[2 + c] for c in range(col_start - 2, col_start - 2 + span))
                tk.Label(inner, text=label_text, bg=_bg_header, font=("TkDefaultFont", 9, "bold"),
                         width=group_width, anchor="center").grid(row=0, column=col_start,
                                                                 columnspan=span, sticky="nsew", padx=1, pady=1)
            col_start += span

        # Header row 1: Include, Sample, Alias, element names
        tk.Label(inner, text="✓", bg=_bg_header, font=("TkDefaultFont", 11, "bold"), width=col_widths[0], anchor="center").grid(row=1, column=0, sticky="nsew", padx=1, pady=1)
        tk.Label(inner, text="Sample", bg=_bg_header, font=("TkDefaultFont", 11, "bold"), width=col_widths[1], anchor="w").grid(row=1, column=1, sticky="nsew", padx=1, pady=1)
        tk.Label(inner, text="Alias", bg=_bg_header, font=("TkDefaultFont", 11, "bold"), width=col_widths[2], anchor="w").grid(row=1, column=2, sticky="nsew", padx=1, pady=1)
        for c, (element, unit_type) in enumerate(progress_cols):
            tk.Label(inner, text=element, bg=_bg_header, font=("TkDefaultFont", 11, "bold"),
                     width=col_widths[3 + c], anchor="center").grid(row=1, column=3 + c, sticky="nsew", padx=1, pady=1)

        # Data rows: each cell is a Label with its own background
        for r, sample in enumerate(self.progress_samples, start=2):
            incl = self.sample_include.get(sample, True)
            incl_text = "☑" if incl else "☐"
            incl_lbl = tk.Label(inner, text=incl_text, bg=_bg_include_sample, font=("TkDefaultFont", 11), width=col_widths[0], anchor="center", cursor="hand2")
            incl_lbl.grid(row=r, column=0, sticky="nsew", padx=1, pady=1)
            incl_lbl.bind("<ButtonRelease-1>", lambda e, s=sample: self._toggle_sample_include(s))
            # Sample ID (read-only)
            tk.Label(inner, text=sample, bg=_bg_include_sample, font=("TkDefaultFont", 11),
                     width=col_widths[1], anchor="w").grid(row=r, column=1, sticky="nsew", padx=1, pady=1)
            # Alias (clickable to edit)
            alias_text = self.sample_aliases.get(sample, "")
            alias_lbl = tk.Label(
                inner,
                text=alias_text,
                bg=_bg_include_sample,
                font=("TkDefaultFont", 11, "italic"),
                width=col_widths[2],
                anchor="w",
                cursor="xterm",
            )
            alias_lbl.grid(row=r, column=2, sticky="nsew", padx=1, pady=1)
            alias_lbl.bind("<ButtonRelease-1>", lambda e, s=sample: self._edit_sample_alias(s))

            for c, (element, unit_type) in enumerate(progress_cols):
                status = self.progress_data.get((sample, element, unit_type))
                if status == 'complete':
                    bg, text = _bg_complete, "✓"
                elif status == 'partial':
                    bg, text = _bg_partial, ""  # Partial: colour only, no symbol
                elif status == 'missing_file':
                    bg, text = _bg_missing_file, "!"  # No input file
                else:
                    bg, text = _bg_missing, ""  # Not started
                tk.Label(inner, text=text, bg=bg, font=("TkDefaultFont", 11),
                         width=col_widths[3 + c], anchor="center").grid(row=r, column=3 + c, sticky="nsew", padx=1, pady=1)

        inner.update_idletasks()
        if hasattr(self, 'progress_table_canvas'):
            self.progress_table_canvas.configure(scrollregion=self.progress_table_canvas.bbox("all"))

    def update_sample_element_progress(self, sample, element, status='complete'):
        """Update progress for a specific sample-element pair. Updates all (sample, element, unit_type) entries."""
        for (s, e, ut), current in list(self.progress_data.items()):
            if s == sample and e == element:
                if current == 'missing_file':
                    continue
                self.progress_data[(s, e, ut)] = status
        if hasattr(self, 'progress_table') and self.progress_table is not None:
            self.update_progress_table()

    def _edit_sample_alias(self, sample):
        """Prompt user to edit the display alias for a sample."""
        current = self.sample_aliases.get(sample, "")
        initial = current or sample
        new_alias = simpledialog.askstring("Rename sample", f"Alias for sample '{sample}':", initialvalue=initial)
        if new_alias is None:
            return  # User cancelled
        new_alias = new_alias.strip()
        if not new_alias or new_alias == sample:
            # Clear alias if empty or identical to sample ID
            if sample in self.sample_aliases:
                del self.sample_aliases[sample]
        else:
            self.sample_aliases[sample] = new_alias
        # Persist aliases and refresh table
        self._save_sample_aliases()
        if hasattr(self, 'progress_table') and self.progress_table is not None:
            self.update_progress_table()
    
    def check_sample_element_status(self, sample, element, unit_type='ppm'):
        """Check the current status of a sample-element-unit based on output files."""
        if not self.output_dir:
            return None
        
        elem_out = f"{element}_{unit_type}"
        element_dir = os.path.join(self.output_dir, elem_out)
        # Check if composite exists (complete)
        composite_path = os.path.join(element_dir, f"{elem_out}_composite.png")
        if os.path.exists(composite_path):
            return 'complete'
        
        # Check if histogram exists (partial)
        hist_path = os.path.join(element_dir, 'Histograms', f"{sample}_histogram.png")
        if os.path.exists(hist_path):
            return 'partial'
        
        return None

    def batch_process_all_elements(self):
        """Process all elements found in the input folder automatically."""
        if not self.input_dir:
            custom_dialogs.showerror(self.master, "Error", "Please select an input folder first.")
            return
        
        # Only prompt for output folder if not already selected
        if not self.output_dir or not os.path.isdir(self.output_dir):
            self.select_output_folder()
            if not self.output_dir or not os.path.isdir(self.output_dir):
                return  # User cancelled
        
        # Use progress_columns (element, unit) so ppm and CPS are processed separately
        self.scan_progress_table()
        pairs = getattr(self, 'progress_columns', [])
        if not pairs:
            custom_dialogs.showwarning(self.master, "No Elements", "No element files found in the input folder.")
            return
        
        # Ask for confirmation
        num_pairs = len(pairs)
        pair_labels = [f"{e} ({u})" for e, u in pairs]
        result = custom_dialogs.askyesno(
            self.master, "Batch Process Confirmation",
            f"This will process {num_pairs} element-unit combination(s):\n\n" +
            ", ".join(pair_labels) +
            "\n\nEach will:\n" +
            "1. Load data and generate histograms/statistics\n" +
            "2. Use the suggested 99th percentile as scale max\n" +
            "3. Generate and save the composite\n\n" +
            "This may take a while. Continue?"
        )
        
        if not result:
            return
        
        # Show progress bar and disable batch button so UI updates during run
        self._batch_running = True
        if hasattr(self, 'batch_btn'):
            self.batch_btn.config(state=tk.DISABLED)
        if hasattr(self, 'batch_progress_bar'):
            self.batch_progress_bar['value'] = 0
            self.batch_progress_bar['maximum'] = 100
        if hasattr(self, 'batch_progress_label'):
            self.batch_progress_label.config(text="Starting...")
        self.master.update()
        
        try:
            # Process each (element, unit) pair
            self.log_print(f"\n🚀 Starting batch processing of {num_pairs} element-unit combination(s)...")
            successful = 0
            failed = []
            
            for idx, (elem, ut) in enumerate(pairs, 1):
                pair_label = f"{elem} ({ut})"
                try:
                    # Update progress bar and label so user sees live progress
                    if hasattr(self, 'batch_progress_bar'):
                        self.batch_progress_bar['value'] = (idx - 1) / num_pairs * 100
                    if hasattr(self, 'batch_progress_label'):
                        self.batch_progress_label.config(text=f"{idx}/{num_pairs}: {pair_label}")
                    self.master.update()
                    
                    self.log_print(f"\n[{idx}/{num_pairs}] Processing {pair_label}...")
                    
                    # Set element and unit
                    self.element.set(elem)
                    self.unit.set(ut)
                    self.master.update()
                    
                    # Load data (this sets scale_max automatically to 99th percentile)
                    self.load_data()
                    
                    # Check if data was loaded successfully
                    if not self.matrices:
                        self.log_print(f"⚠️  No data loaded for {pair_label}, skipping...")
                        failed.append((pair_label, "No data found"))
                        continue
                    
                    # Generate and save composite directly (no preview); use best layout (no extra blank rows)
                    self.log_print(f"  Generating composite for {pair_label}...")
                    best_rows = self._best_composite_rows(len(self.matrices))
                    self.generate_composite(preview=False, override_rows=best_rows)
                    
                    successful += 1
                    self.log_print(f"✅ Completed {pair_label} ({idx}/{num_pairs})")
                    
                    # Move progress bar to completed
                    if hasattr(self, 'batch_progress_bar'):
                        self.batch_progress_bar['value'] = idx / num_pairs * 100
                    self.master.update()
                    
                except Exception as e:
                    error_msg = str(e)
                    self.log_print(f"❌ Error processing {pair_label}: {error_msg}")
                    failed.append((pair_label, error_msg))
                    import traceback
                    self.log_print(traceback.format_exc())
                    if hasattr(self, 'batch_progress_bar'):
                        self.batch_progress_bar['value'] = idx / num_pairs * 100
                    self.master.update()
                    continue

            # Summary (after loop completes)
            self.log_print(f"\n{'='*50}")
            self.log_print(f"Batch processing complete!")
            self.log_print(f"✅ Successful: {successful}/{num_pairs}")
            if failed:
                self.log_print(f"❌ Failed: {len(failed)}")
                for pair_label, reason in failed:
                    self.log_print(f"   - {pair_label}: {reason}")
            self.log_print(f"{'='*50}")
        finally:
            self._batch_running = False
            if hasattr(self, 'batch_btn'):
                self.batch_btn.config(state=tk.NORMAL)
            if hasattr(self, 'batch_progress_bar'):
                self.batch_progress_bar['value'] = 100
            if hasattr(self, 'batch_progress_label'):
                self.batch_progress_label.config(text="Complete")
            self.master.update()
        
        # Show completion message
        if failed:
            custom_dialogs.showwarning(
                self.master, "Batch Processing Complete",
                f"Processed {successful}/{num_pairs} element-unit combination(s) successfully.\n\n"
                f"{len(failed)} failed. Check the log for details."
            )
        else:
            custom_dialogs.showinfo(
                self.master, "Batch Processing Complete",
                f"Successfully processed all {num_pairs} element-unit combination(s)!"
            )

    def update_element_dropdown(self):
        elements = set()
        for file in glob.glob(os.path.join(self.input_dir, "* matrix.xlsx")):
            parsed = self.parse_matrix_filename(file)
            if parsed:
                _, element, _ = parsed
                elements.add(element)
        
        if elements:
            element_list = sorted(list(elements))
            # Update both dropdowns if they exist
            if hasattr(self, 'element_dropdown'):
                self.element_dropdown['values'] = element_list
                self.element_dropdown['state'] = 'readonly'
            if hasattr(self, 'element_dropdown_preview'):
                self.element_dropdown_preview['values'] = element_list
                self.element_dropdown_preview['state'] = 'readonly'
            # Set the first element as default if no element is currently selected
            if not self.element.get() or self.element.get() not in element_list:
                self.element.set(next(iter(elements)))
            self.update_unit_dropdown()
        else:
            self.log_print("No valid element files found in the selected directory.")

    def get_element_output_subdir(self):
        """Return output subdirectory for current element+unit, e.g. 'Fe_ppm' or 'Ca_raw'."""
        element = self.element.get()
        unit = self.unit.get()
        if not element:
            return ""
        if unit:
            return f"{element}_{unit}"
        return element  # fallback for backward compat

    def update_unit_dropdown(self):
        """Populate unit dropdown with units that exist for the selected element."""
        element = self.element.get()
        units = []
        for (e, u) in getattr(self, 'progress_columns', []):
            if e == element:
                units.append(u)
        # Fallback: scan files if progress_columns not populated
        if not units and self.input_dir and os.path.isdir(self.input_dir):
            seen = set()
            for file in glob.glob(os.path.join(self.input_dir, f"* {element}_* matrix.xlsx")):
                parsed = self.parse_matrix_filename(file)
                if parsed:
                    _, _, ut = parsed
                    seen.add(ut)
            for file in glob.glob(os.path.join(self.input_dir, f"* {element} matrix.xlsx")):
                parsed = self.parse_matrix_filename(file)
                if parsed:
                    _, _, ut = parsed
                    seen.add(ut)
            units = sorted(seen, key=lambda u: ('ppm', 'CPS', 'raw').index(u) if u in ('ppm', 'CPS', 'raw') else 99)
        if units:
            for w in [getattr(self, 'unit_dropdown', None), getattr(self, 'unit_dropdown_preview', None)]:
                if w and w.winfo_exists():
                    w['values'] = units
                    w['state'] = 'readonly'
            if not self.unit.get() or self.unit.get() not in units:
                self.unit.set(units[0])
        else:
            for w in [getattr(self, 'unit_dropdown', None), getattr(self, 'unit_dropdown_preview', None)]:
                if w and w.winfo_exists():
                    w['values'] = []
                    w['state'] = 'readonly'
            self.unit.set("")

    def on_resize(self, event):
        """Handle resize events for preview container (legacy - preview now in separate window)."""
        # Only update if preview_image exists and we're still using main window preview
        if hasattr(self, 'preview_image') and self.preview_image is not None:
            # Only update if the resize is significant enough to avoid excessive updates
            if hasattr(self, '_last_resize_time'):
                current_time = time.time()
                if current_time - self._last_resize_time < 0.1:  # Throttle to max 10 updates per second
                    return
                self._last_resize_time = current_time
            else:
                self._last_resize_time = time.time()
            
            self.update_preview_image()

    def log_print(self, message, status_only=False):
        """Print message to log. If status_only=True, only show status-related messages."""
        # Helper function to write to a log widget
        def write_to_log(log_widget):
            if log_widget:
                if status_only:
                    # Only log status changes and important messages
                    if any(keyword in message.lower() for keyword in ['status', 'idle', 'busy', 'finishing', 'complete', 'error', 'warning', '⚠️', '✅', '❌']):
                        log_widget.insert(tk.END, message + "\n")
                        log_widget.see(tk.END)
                else:
                    # Show all messages for now, but we'll filter specific ones
                    # Skip verbose debug and progress messages
                    skip_keywords = ['found composite', 'skipping subplot', 'debug:', 'figure created', 'grid layout']
                    if not any(keyword in message.lower() for keyword in skip_keywords):
                        log_widget.insert(tk.END, message + "\n")
                        log_widget.see(tk.END)
        
        # Write to both log widgets (main log and preview tab log)
        write_to_log(self.log if hasattr(self, 'log') else None)
        write_to_log(self.log_preview if hasattr(self, 'log_preview') else None)
        
        # During batch or Calculate Statistics, process events so Status Log and progress bar update in real time
        if getattr(self, '_batch_running', False) or getattr(self, '_stats_calculating', False):
            self.master.update()
        else:
            self.master.update_idletasks()
    
    def set_status(self, status):
        """Set the application status (Idle, Busy, Finishing)."""
        self.status = status
        self.log_print(f"Status: {status}", status_only=True)

    def load_matrix_2d(self, path):
        """Load a 2D matrix from an Excel file, with error handling for Dropbox sync issues."""
        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            ws = wb.active
            return np.array([[cell if isinstance(cell, (int, float)) and cell >= 0 else np.nan for cell in row] for row in ws.iter_rows(values_only=True)])
        except KeyError as e:
            # This often happens with Dropbox placeholder files that aren't fully synced
            error_msg = str(e)
            if "Content_Types" in error_msg or "archive" in error_msg.lower():
                raise FileNotFoundError(
                    f"File appears to be incomplete or not fully synced (Dropbox placeholder?): {os.path.basename(path)}\n"
                    f"Please ensure the file is fully downloaded before loading."
                )
            else:
                raise
        except Exception as e:
            # Re-raise with more context
            raise Exception(f"Error loading file {os.path.basename(path)}: {str(e)}")

    def downsample_matrix(self, matrix, target_max=512):
        """Downsample matrix for faster preview rendering."""
        h, w = matrix.shape
        scale = max(h, w) / target_max
        if scale <= 1:
            return matrix
        
        # Calculate downsampling factors
        sh = max(1, int(scale))
        sw = max(1, int(scale))
        
        # Ensure dimensions are divisible by downsampling factors
        new_h = (h // sh) * sh
        new_w = (w // sw) * sw
        
        # Crop to divisible dimensions
        cropped = matrix[:new_h, :new_w]
        
        # Reshape and average
        reshaped = cropped.reshape(new_h//sh, sh, new_w//sw, sw)
        downsampled = reshaped.mean(axis=(1, 3))
        
        return downsampled

    def import_custom_pixel_sizes(self):
        custom_dialogs.showinfo(self.master, "Load Custom Physical Pixel Size", "Select CSV file with custom pixel sizes (Cancel to generate template)")

        file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")], title="Select Custom Pixel Sizes CSV (Cancel to generate template)")
        if file_path:
            try:
                df = pd.read_csv(file_path)
                self.custom_pixel_sizes = dict(zip(df['Sample'], df['Pixel Size']))
                self.log_print(f"Imported custom pixel sizes for {len(self.custom_pixel_sizes)} samples.")
            except Exception as e:
                self.log_print(f"Error importing custom pixel sizes: {str(e)}")
                self.generate_pixel_size_template()
        else:
            self.generate_pixel_size_template()

    def generate_pixel_size_template(self):
        if not self.input_dir:
            custom_dialogs.showerror(self.master, "Error", "Please select an input folder first.")
            return

        samples = set()
        for file in glob.glob(os.path.join(self.input_dir, "* matrix.xlsx")):
            parsed = self.parse_matrix_filename(file)
            if parsed:
                sample, _, _ = parsed
                samples.add(sample)

        if not samples:
            custom_dialogs.showerror(self.master, "Error", "No valid sample files found in the input directory.")
            return

        df = pd.DataFrame({'Sample': list(samples), 'Pixel Size': self.pixel_size.get()})
        
        save_path = filedialog.asksaveasfilename(
            title="Save Pixel Size Template",
            defaultextension=".csv", 
            filetypes=[("CSV files", "*.csv")]
        )
        if save_path:
            df.to_csv(save_path, index=False)
            self.log_print(f"✅ Saved pixel size template to {save_path}")

    def handle_pixel_sizes(self):
        if not self.input_dir:
            custom_dialogs.showerror(self.master, "Error", "Please select an input folder first.")
            return

        self.import_custom_pixel_sizes()

    def load_data(self):
        if not self.input_dir:
            custom_dialogs.showerror(self.master, "Error", "Please select an input folder first.")
            return
        # Only prompt for output folder if not already selected
        if not self.output_dir or not os.path.isdir(self.output_dir):
            self.select_output_folder()

        self.set_status("Busy")
        element = self.element.get()
        unit = self.unit.get()
        # Search for files with old format (ppm/CPS) and new format (raw, no unit)
        # Filter by selected unit so ppm and CPS are never mixed
        if unit == 'ppm':
            pattern = os.path.join(self.input_dir, f"* {element}_ppm matrix.xlsx")
            files = sorted(glob.glob(pattern))
        elif unit == 'CPS':
            pattern = os.path.join(self.input_dir, f"* {element}_CPS matrix.xlsx")
            files = sorted(glob.glob(pattern))
        elif unit == 'raw':
            pattern = os.path.join(self.input_dir, f"* {element} matrix.xlsx")
            files = [f for f in glob.glob(pattern) if (p := self.parse_matrix_filename(f)) and p[2] == 'raw']
        else:
            # No unit selected - check if both ppm and CPS exist (would mix units)
            pattern_ppm = os.path.join(self.input_dir, f"* {element}_ppm matrix.xlsx")
            pattern_CPS = os.path.join(self.input_dir, f"* {element}_CPS matrix.xlsx")
            has_ppm = bool(glob.glob(pattern_ppm))
            has_CPS = bool(glob.glob(pattern_CPS))
            if has_ppm and has_CPS:
                custom_dialogs.showerror(self.master, "Unit Required",
                    f"Element {element} has both ppm and CPS files. Please select a unit from the Unit dropdown.")
                return
            # Single unit or raw only - use appropriate pattern
            if has_ppm:
                files = sorted(glob.glob(pattern_ppm))
            elif has_CPS:
                files = sorted(glob.glob(pattern_CPS))
            else:
                pattern_raw = os.path.join(self.input_dir, f"* {element} matrix.xlsx")
                files = [f for f in glob.glob(pattern_raw) if (p := self.parse_matrix_filename(f)) and p[2] == 'raw']
        # Restrict to selected samples (Progress table Include column)
        selected = set(self.get_selected_samples())
        files = [f for f in files if (parsed := self.parse_matrix_filename(f)) and parsed[0] in selected and (not unit or parsed[2] == unit)]

        if not files:
            custom_dialogs.showerror(self.master, "Error", f"No files for element {element} among selected samples. Use the Progress table checkboxes to include at least one sample with data for this element.")
            return

        # Share progress bar with Batch so user sees live progress (same as batch processing)
        num_files = len(files)
        self._stats_calculating = True
        if hasattr(self, 'summarize_btn'):
            self.summarize_btn.config(state=tk.DISABLED)
        if hasattr(self, 'batch_progress_bar'):
            self.batch_progress_bar['value'] = 0
            self.batch_progress_bar['maximum'] = 100
        if hasattr(self, 'batch_progress_label'):
            self.batch_progress_label.config(text="Calculate statistics: Starting...")
        self.master.update()

        try:
            # Check for existing statistics.csv to enable incremental processing
            elem_out = self.get_element_output_subdir()
            stats_path = os.path.join(self.output_dir, elem_out, f"{elem_out}_statistics.csv")
            existing_stats_df = None
            existing_samples = set()

            if os.path.exists(stats_path):
                try:
                    existing_stats_df = pd.read_csv(stats_path)
                    if 'Sample' in existing_stats_df.columns:
                        existing_samples = set(existing_stats_df['Sample'].tolist())
                        # Progress table will show this information
                except Exception as e:
                    self.log_print(f"⚠️ Could not read existing statistics: {e}")

            # Determine if we need to calculate statistics
            # Get list of samples from files
            samples_from_files = set()
            for f in files:
                parsed = self.parse_matrix_filename(f)
                if parsed:
                    sample, parsed_element, _ = parsed
                    if parsed_element == element:
                        samples_from_files.add(sample)

            # Check if all samples already have statistics
            all_samples_exist = existing_samples and samples_from_files.issubset(existing_samples)

            if all_samples_exist:
                self.log_print("Status: Busy - Loading matrices (statistics already exist, skipping calculation)...")
            else:
                self.log_print("Status: Busy - Loading matrices and calculating statistics...")

            self.matrices = []
            self.labels = []
            self.current_element_unit = None  # ppm, CPS, or raw (for color bar label)
            self.pixel_sizes_by_sample = {}
            percentiles = []
            iqrs = []
            means = []
            new_samples = []  # Track which samples are new

            # If using custom pixel sizes, only load data for samples present in the custom file
            samples_to_load = set(self.custom_pixel_sizes.keys()) if self.use_custom_pixel_sizes.get() else None

            for idx, f in enumerate(files, 1):
                display_name = os.path.basename(f)
                parsed = self.parse_matrix_filename(f)
                if parsed:
                    sample, parsed_element, unit_type = parsed
                    display_name = sample
                    # Verify the element matches (safety check)
                    if parsed_element != element:
                        continue
                    if samples_to_load is None or sample in samples_to_load:
                        if self.current_element_unit is None:
                            self.current_element_unit = unit_type
                        # Check if this sample is new
                        is_new = sample not in existing_samples

                        try:
                            matrix = self.load_matrix_2d(f)
                            self.labels.append(sample)
                            self.matrices.append(matrix)
                        except (FileNotFoundError, Exception) as e:
                            # Handle Dropbox sync issues or other file loading errors
                            error_msg = str(e)
                            self.log_print(f"❌ Failed to load {sample}: {error_msg}")
                            # Continue processing other files instead of stopping
                            continue

                        # Get the pixel size for this sample
                        if self.use_custom_pixel_sizes.get() and sample in self.custom_pixel_sizes:
                            pixel_size = self.custom_pixel_sizes[sample]
                        else:
                            pixel_size = self.pixel_size.get()

                        self.pixel_sizes_by_sample[sample] = pixel_size

                        # Only process statistics and histograms for new samples
                        if is_new:
                            new_samples.append(sample)

                            # Calculate percentiles, IQR, and mean
                            p25, p50, p75, p99 = np.nanpercentile(matrix, [25, 50, 75, 99])
                            iqr = p75 - p25
                            mean = np.nanmean(matrix)
                            percentiles.append((sample, p25, p50, p75, p99))
                            iqrs.append((sample, iqr))
                            means.append((sample, mean))

                            # Generate and save histogram
                            plt.figure(figsize=(10, 6))
                            plt.hist(matrix.flatten(), bins=50, range=(0, np.nanpercentile(matrix, 99)))
                            plt.title(f"Histogram for {sample}")
                            plt.xlabel("Value")
                            plt.ylabel("Frequency")
                            hist_path = os.path.join(self.output_dir, elem_out, 'Histograms', f"{sample}_histogram.png")
                            os.makedirs(os.path.dirname(hist_path), exist_ok=True)
                            plt.savefig(hist_path)
                            plt.close()

                            # Update progress table for this sample
                            if hasattr(self, 'progress_table') and self.progress_table:
                                self.update_sample_element_progress(sample, element, 'partial')
                                self.update_progress_table()

                # Update shared progress bar every file so user sees live progress (same as batch)
                if hasattr(self, 'batch_progress_bar'):
                    self.batch_progress_bar['value'] = (idx / num_files) * 100
                if hasattr(self, 'batch_progress_label'):
                    self.batch_progress_label.config(text=f"Sample {idx} of {num_files}: {display_name}")
                self.master.update()

            # Merge new statistics with existing ones
            if existing_stats_df is not None and new_samples:
                # Create new statistics DataFrame
                new_percentiles_df = pd.DataFrame(percentiles, columns=['Sample', '25th Percentile', '50th Percentile', '75th Percentile', '99th Percentile'])
                new_iqr_df = pd.DataFrame(iqrs, columns=['Sample', 'IQR'])
                new_mean_df = pd.DataFrame(means, columns=['Sample', 'Mean'])
                new_stats_df = new_percentiles_df.merge(new_iqr_df, on='Sample').merge(new_mean_df, on='Sample')
                # Attach aliases for new samples
                new_stats_df['Alias'] = new_stats_df['Sample'].map(lambda s: self.sample_aliases.get(s, s))

                # Ensure existing stats have Alias column as well
                if 'Alias' not in existing_stats_df.columns:
                    existing_stats_df['Alias'] = existing_stats_df['Sample'].map(lambda s: self.sample_aliases.get(s, s))

                # Combine with existing
                stats_df = pd.concat([existing_stats_df, new_stats_df], ignore_index=True)

                # Round statistics and save
                stats_df = stats_df.map(lambda x: float(f"{x:.5g}") if isinstance(x, (int, float)) else x)
                os.makedirs(os.path.dirname(stats_path), exist_ok=True)
                stats_df.to_csv(stats_path, index=False)
            elif existing_stats_df is not None:
                # No new samples, use existing (don't need to recalculate or save)
                stats_df = existing_stats_df
                self.log_print(f"✓ Using existing statistics for {len(existing_samples)} sample(s)")
            else:
                # No existing stats, create new
                percentiles_df = pd.DataFrame(percentiles, columns=['Sample', '25th Percentile', '50th Percentile', '75th Percentile', '99th Percentile'])
                iqr_df = pd.DataFrame(iqrs, columns=['Sample', 'IQR'])
                mean_df = pd.DataFrame(means, columns=['Sample', 'Mean'])
                stats_df = percentiles_df.merge(iqr_df, on='Sample').merge(mean_df, on='Sample')
                # Attach aliases column
                stats_df['Alias'] = stats_df['Sample'].map(lambda s: self.sample_aliases.get(s, s))

                # Round statistics and save
                stats_df = stats_df.map(lambda x: float(f"{x:.5g}") if isinstance(x, (int, float)) else x)
                os.makedirs(os.path.dirname(stats_path), exist_ok=True)
                stats_df.to_csv(stats_path, index=False)

            # Update statistics table display
            self.update_statistics_table(stats_df)

            # Update progress table
            if hasattr(self, 'progress_table') and self.progress_table:
                self._check_existing_progress()
                self.update_progress_table()

            self.set_status("Idle")
            self.log_print("Status: Idle - Statistics calculation complete.")

            # Set scale_max based on 99th percentile of ALL data (existing + new)
            overall_99th = np.nanpercentile(np.hstack([m.flatten() for m in self.matrices]), 99)
            self.scale_max.set(round(overall_99th,3))
            self.log_print(f"Scale max set to {self.scale_max.get():.2f} based on overall 99th percentile (all {len(self.matrices)} sample(s))")

            # Update progress table - mark as partial (histograms and stats done)
            element = self.element.get()
            for sample in self.labels:
                self.update_sample_element_progress(sample, element, 'partial')

            # Update progress table display
            if hasattr(self, 'progress_table') and self.progress_table:
                self._check_existing_progress()
                self.update_progress_table()

            self._update_save_matrix_button_state()
        finally:
            self._stats_calculating = False
            if hasattr(self, 'summarize_btn'):
                self.summarize_btn.config(state=tk.NORMAL)
            if hasattr(self, 'batch_progress_bar'):
                self.batch_progress_bar['value'] = 100
            if hasattr(self, 'batch_progress_label'):
                self.batch_progress_label.config(text="Complete")

    def _sample_aliases_path(self):
        """Return path to the CSV that stores global sample aliases for this output directory."""
        return os.path.join(self.output_dir, "sample_aliases.csv")

    def _load_sample_aliases(self):
        """Load sample aliases from CSV if present."""
        path = self._sample_aliases_path()
        if not path or not os.path.exists(path):
            return
        try:
            df = pd.read_csv(path)
            if 'Sample' in df.columns and 'Alias' in df.columns:
                for _, row in df[['Sample', 'Alias']].dropna().iterrows():
                    sample = str(row['Sample'])
                    alias = str(row['Alias']).strip()
                    if alias and alias != sample:
                        self.sample_aliases[sample] = alias
        except Exception as e:
            self.log_print(f"⚠️ Could not load sample aliases: {e}")

    def _save_sample_aliases(self):
        """Persist current sample aliases to CSV."""
        path = self._sample_aliases_path()
        if not path:
            return
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            if not self.sample_aliases:
                # If there are no aliases, avoid writing an empty file
                if os.path.exists(path):
                    try:
                        os.remove(path)
                    except Exception:
                        pass
                return
            data = [{'Sample': s, 'Alias': a} for s, a in sorted(self.sample_aliases.items()) if a]
            if not data:
                return
            df = pd.DataFrame(data)
            df.to_csv(path, index=False)
        except Exception as e:
            self.log_print(f"⚠️ Could not save sample aliases: {e}")

    def show_preview_window(self):
        """Display the preview image in a separate window."""
        if not hasattr(self, 'preview_image') or self.preview_image is None:
            return
        
        # Close existing preview window if open
        if self.preview_window is not None and self.preview_window.winfo_exists():
            self.preview_window.destroy()
        
        # Create new preview window
        self.preview_window = tk.Toplevel(self.master)
        element = self.element.get()
        self.preview_window.title(f"Composite Preview - {element}")
        
        # Make window resizable
        self.preview_window.resizable(True, True)
        
        # Get image dimensions
        img_width, img_height = self.preview_image.size
        
        # Set window size to fit image (with some padding), but cap at screen size
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        
        # Scale down if image is larger than 90% of screen
        max_width = int(screen_width * 0.9)
        max_height = int(screen_height * 0.9)
        
        if img_width > max_width or img_height > max_height:
            # Calculate scaling factor
            scale_w = max_width / img_width
            scale_h = max_height / img_height
            scale = min(scale_w, scale_h)
            display_width = int(img_width * scale)
            display_height = int(img_height * scale)
        else:
            display_width = img_width
            display_height = img_height
        
        # Set window geometry (add padding for window chrome and control panel)
        control_panel_height = 100  # Space for controls at bottom
        window_width = display_width + 20
        window_height = display_height + 40 + control_panel_height
        
        # Set minimum size to ensure controls are always visible
        self.preview_window.minsize(400, 300)
        
        # Set initial geometry
        self.preview_window.geometry(f"{window_width}x{window_height}")
        
        # Center the window on screen
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.preview_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Create frame for the image
        preview_frame = tk.Frame(self.preview_window)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create label to display image (store reference for updates)
        # Start with empty label, will be populated after window is laid out
        self.preview_window_label = tk.Label(preview_frame)
        self.preview_window_label.pack(fill=tk.BOTH, expand=True)
        
        # Add info label
        info_label = tk.Label(preview_frame, 
                             text=f"{element} Composite - {img_width}x{img_height} pixels",
                             font=("Arial", 9), foreground="gray")
        info_label.pack(pady=(5, 0))
        
        # Create minimal control panel at bottom
        control_panel = tk.Frame(self.preview_window, bg='#f0f0f0', relief=tk.RAISED, borderwidth=1)
        control_panel.pack(fill=tk.X, side=tk.BOTTOM, padx=0, pady=0)
        
        # Left side: show current font size (set in Fonts group)
        font_frame = tk.Frame(control_panel, bg='#f0f0f0')
        font_frame.pack(side=tk.LEFT, padx=10, pady=8)
        tk.Label(font_frame, text="Element label (Label font size):", font=("Arial", 9), bg='#f0f0f0').pack(side=tk.LEFT, padx=(0, 5))
        font_value_label = tk.Label(font_frame, text=f"{self._pt_from_font(self.element_label_font, 16)} pt", font=("Arial", 9), bg='#f0f0f0', width=4)
        font_value_label.pack(side=tk.LEFT)
        
        # Right side: Save button
        button_frame = tk.Frame(control_panel, bg='#f0f0f0')
        button_frame.pack(side=tk.RIGHT, padx=10, pady=8)
        
        # Save Composite button with icon
        save_icon = self.button_icons.get('save')
        if save_icon:
            save_btn = tk.Button(button_frame, image=save_icon, 
                               command=self.save_composite,
                               padx=2, pady=8, bg="#4CAF50", fg="black", relief='raised',
                               activebackground='#45a049')
            save_btn.image = save_icon  # Keep reference
        else:
            save_btn = tk.Button(button_frame, text="💾", 
                                command=self.save_composite,
                                padx=2, pady=8, bg="#4CAF50", fg="black", relief='raised')
        save_btn.pack(side=tk.LEFT)
        
        # Handle window close
        self.preview_window.protocol("WM_DELETE_WINDOW", self._close_preview_window)
        
        # Bind resize event to update image size
        self.preview_window.bind("<Configure>", self._on_preview_window_resize)
        
        # Force window update and then set initial image size after layout is complete
        self.preview_window.update_idletasks()
        self.preview_window.after(10, self._update_preview_window_image)
    
    def _on_preview_window_resize(self, event):
        """Handle resize events for the preview window."""
        # Only update if this is the preview window being resized (not other widgets)
        if event.widget != self.preview_window:
            return
        
        # Throttle resize updates to avoid excessive redraws
        if not hasattr(self, '_preview_resize_time'):
            self._preview_resize_time = time.time()
        
        current_time = time.time()
        if current_time - self._preview_resize_time < 0.1:  # Max 10 updates per second
            return
        self._preview_resize_time = current_time
        
        # Update the image to fit the new window size
        self._update_preview_window_image()
    
    def _update_preview_window_image(self):
        """Update the image displayed in the preview window without recreating the window."""
        if not hasattr(self, 'preview_window_label') or self.preview_window_label is None:
            return
        
        if not hasattr(self, 'preview_image') or self.preview_image is None:
            return
        
        # Get available space from the preview frame (parent of the label)
        if hasattr(self, 'preview_window') and self.preview_window.winfo_exists():
            # Get the frame that contains the label
            preview_frame = self.preview_window_label.master
            # Force update to get accurate dimensions
            preview_frame.update_idletasks()
            frame_width = preview_frame.winfo_width()
            frame_height = preview_frame.winfo_height()
            
            # Account for padding (10px on each side) and info label space (~30px)
            available_width = max(100, frame_width - 20)  # Subtract padding
            available_height = max(100, frame_height - 50)  # Subtract padding and info label space
        else:
            available_width = 400
            available_height = 300
        
        # Get original image dimensions
        img_width, img_height = self.preview_image.size
        
        # Calculate aspect ratios
        img_aspect = img_width / img_height
        available_aspect = available_width / available_height
        
        # Resize to fit available space while maintaining aspect ratio
        if available_aspect > img_aspect:
            # Available space is wider - fit to height
            display_height = available_height
            display_width = int(display_height * img_aspect)
        else:
            # Available space is taller - fit to width
            display_width = available_width
            display_height = int(display_width / img_aspect)
        
        # Ensure minimum size
        display_width = max(100, display_width)
        display_height = max(100, display_height)
        
        # Resize and update image
        display_image = self.preview_image.resize((display_width, display_height), Image.LANCZOS)
        tk_image = ImageTk.PhotoImage(display_image)
        self.preview_window_label.config(image=tk_image)
        self.preview_window_label.image = tk_image  # Keep reference
    
    def _close_preview_window(self):
        """Close the preview window."""
        if self.preview_window:
            self.preview_window.destroy()
            self.preview_window = None
            self.preview_window_label = None
    
    def preview_composite(self):
        try:
            self.generate_composite(preview=True)
            # Track which element+unit was last previewed for safety validation
            self._last_previewed_element_unit = self.get_element_output_subdir()
            self.set_status("Idle")
            self.log_print("Status: Idle - Preview generated.")
        except Exception as e:
            self.set_status("Idle")
            self.log_print(f"❌ Error generating preview: {e}")

    def _update_save_matrix_button_state(self):
        """Disable Save Composite Matrix when no data or when multiple pixel sizes (saved file has no per-panel scale)."""
        if not hasattr(self, 'save_matrix_btn') or not self.save_matrix_btn.winfo_exists():
            return
        no_data = not getattr(self, 'matrices', None) or len(self.matrices) == 0
        multiple_sizes = self.use_custom_pixel_sizes.get()
        self.save_matrix_btn.config(state=tk.DISABLED if (no_data or multiple_sizes) else tk.NORMAL)

    def save_composite_matrix(self):
        """Save a composite matrix file that can be opened in muaddata for polygon selection.
        Output is a grid of matrices (with NaN padding/separators); pixel size is not stored in the file."""
        if not self.matrices:
            custom_dialogs.showerror(self.master, "Error", "No data loaded. Please load data first.")
            return
        
        self.set_status("Busy")
        self.log_print("Status: Busy - Creating composite matrix...")
        
        element = self.element.get()
        n = len(self.matrices)
        if getattr(self, 'use_best_layout', None) and self.use_best_layout.get():
            rows = self._best_composite_rows(n)
        else:
            rows = min(self.num_rows.get(), n)
        cols = math.ceil(n / rows)
        
        # Find maximum dimensions
        max_height = max(m.shape[0] for m in self.matrices)
        max_width = max(m.shape[1] for m in self.matrices)
        
        # Pad all matrices to the same size (pad with NaN)
        padded_matrices = []
        for matrix in self.matrices:
            h, w = matrix.shape
            pad_h = max_height - h
            pad_w = max_width - w
            # Pad on bottom and right with NaN
            padded = np.pad(matrix, ((0, pad_h), (0, pad_w)), mode='constant', constant_values=np.nan)
            padded_matrices.append(padded)
        
        # Create separator (row and column of NaN)
        separator_row = np.full((1, max_width), np.nan)
        separator_col = np.full((max_height, 1), np.nan)
        
        # Arrange matrices in grid with separators
        composite_rows = []
        for r in range(rows):
            row_matrices = []
            for c in range(cols):
                idx = r * cols + c
                if idx < len(padded_matrices):
                    row_matrices.append(padded_matrices[idx])
                else:
                    # Empty cell - fill with NaN
                    row_matrices.append(np.full((max_height, max_width), np.nan))
                # Add column separator between matrices (except last)
                if c < cols - 1:
                    row_matrices.append(separator_col)
            
            # Combine matrices in this row horizontally
            if row_matrices:
                row_composite = np.hstack(row_matrices)
                composite_rows.append(row_composite)
                # Add row separator between rows (except last)
                if r < rows - 1:
                    separator_row_full = np.full((1, row_composite.shape[1]), np.nan)
                    composite_rows.append(separator_row_full)
        
        # Combine all rows vertically
        composite_matrix = np.vstack(composite_rows) if composite_rows else np.array([])
        
        # Ask user for save location
        elem_out = self.get_element_output_subdir()
        default_name = f"{elem_out}_composite_matrix.xlsx"
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv")
            ],
            initialfile=default_name,
            initialdir=self.output_dir if self.output_dir else "."
        )
        
        if save_path:
            try:
                if save_path.endswith('.xlsx'):
                    df = pd.DataFrame(composite_matrix)
                    df.to_excel(save_path, header=False, index=False)
                elif save_path.endswith('.csv'):
                    df = pd.DataFrame(composite_matrix)
                    df.to_csv(save_path, header=False, index=False)
                
                self.log_print(f"✓ Composite matrix saved: {os.path.basename(save_path)}")
                self.log_print(f"  Shape: {composite_matrix.shape} (arranged as {rows} rows × {cols} cols)")
                self.log_print(f"  You can now open this file in Muad'Data for polygon selection!")
                custom_dialogs.showinfo(self.master, "Saved", 
                                  f"Composite matrix saved successfully!\n\n"
                                  f"File: {os.path.basename(save_path)}\n"
                                  f"Shape: {composite_matrix.shape}\n"
                                  f"Layout: {rows} rows × {cols} columns\n\n"
                                  f"You can now open this file in Muad'Data's\n"
                                  f"Element Viewer tab for polygon selection.")
                
            except Exception as e:
                custom_dialogs.showerror(self.master, "Save Error", f"Failed to save composite matrix:\n{str(e)}")
                self.log_print(f"❌ Error saving composite matrix: {e}")
        
        self.set_status("Idle")
    
    def save_composite(self):
        self.set_status("Busy")
        self.log_print("Status: Busy - Saving composite...")
        if self.preview_file:
            # Additional safety: Warn user if they're saving a preview that might be from a different element
            elem_out = self.get_element_output_subdir()
            last_key = getattr(self, '_last_previewed_element_unit', None)
            if last_key != elem_out:
                result = custom_dialogs.askyesno(self.master, "Potential Misidentification Warning", 
                    f"Warning: The preview image may be from a different element/unit than the currently selected '{elem_out}'. "
                    f"Do you want to proceed with saving? (It's recommended to generate a new preview first.)")
                if not result:
                    self.log_print("Save cancelled by user due to potential misidentification.")
                    return
            
            out_path = os.path.join(self.output_dir, elem_out, f"{elem_out}_composite.png")
            os.makedirs(os.path.dirname(out_path), exist_ok=True)
            
            # Save the modified preview image (which includes element labels if added)
            if hasattr(self, 'preview_image'):
                self.preview_image.save(out_path, dpi=(300, 300))
            else:
                # Fallback to moving the original file if preview_image is not available
                shutil.move(self.preview_file, out_path)
            
            # Clean up the temporary file
            if os.path.exists(self.preview_file):
                os.remove(self.preview_file)
            self.preview_file = None
            
            # Update progress table - mark as complete
            element = self.element.get()
            for sample in self.labels:
                self.update_sample_element_progress(sample, element, 'complete')
            
            # Update progress table display
            if hasattr(self, 'progress_table') and self.progress_table:
                self._check_existing_progress()
                self.update_progress_table()
            
            self.set_status("Idle")
            self.log_print("Status: Idle - Composite saved.")
        else:
            self.generate_composite(preview=False)

    def _best_composite_rows(self, n):
        """Return the number of rows that minimizes empty cells, then favors a square-ish grid."""
        if n <= 0:
            return 1
        if n == 1:
            return 1
        best_rows = 1
        best_empty = n - 1
        best_diff = n - 1  # |rows - cols|
        for r in range(1, n + 1):
            c = math.ceil(n / r)
            empty = r * c - n
            diff = abs(r - c)
            if empty < best_empty or (empty == best_empty and diff < best_diff):
                best_empty = empty
                best_diff = diff
                best_rows = r
        return best_rows

    def _build_overlay_image(
        self,
        base_image,
        tight_bbox,
        image_positions,
        scale_bar_pos,
        last_ax_width_fig,
        labels,
        show_labels,
        font_size_pt,
        text_color,
        scale_bar_px,
        scale_bar_um,
        scale_bar_caption,
        data_width,
        draw_scale_bar=True,
        scale_bar_font_pt=10,
        draw_element_label=False,
        element_label_text="",
        element_label_font_pt=16,
        draw_credit=False,
        credit_text="",
        credit_font_pt=8,
        dpi=300,
    ):
        """
        Draw a single overlay: sample names, scale bar, and element label in base-image pixel coordinates.
        All annotations share the same layer and the same pt→pixel scaling (font_pt * dpi / 72).
        Returns RGBA PIL Image same size as base_image.
        """
        W, H = base_image.size
        bx0, by0 = getattr(tight_bbox, "x0", 0), getattr(tight_bbox, "y0", 0)
        bx1, by1 = getattr(tight_bbox, "x1", 1), getattr(tight_bbox, "y1", 1)
        bw = bx1 - bx0
        bh = by1 - by0

        def fig_to_pixel(fx, fy):
            px = (fx - bx0) / bw * W if bw else 0
            py = (by0 + bh - fy) / bh * H if bh else 0  # fig y: bottom=0, image y: top=0
            return (int(round(px)), int(round(py)))

        overlay = Image.new("RGBA", (W, H), (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)
        fill = (255, 255, 255) if text_color == "white" else (0, 0, 0)
        font_size_px = max(10, int(font_size_pt * dpi / 72))

        try:
            font = ImageFont.truetype("/System/Library/Fonts/Arial Bold.ttf", font_size_px)
        except Exception:
            try:
                font = ImageFont.truetype("arialbd.ttf", font_size_px)
            except Exception:
                try:
                    font = ImageFont.truetype("arial.ttf", font_size_px)
                except Exception:
                    font = ImageFont.load_default()

        if show_labels and labels and len(image_positions) == len(labels):
            for (x0, y0, w, h), label in zip(image_positions, labels):
                cx_fig = x0 + w * 0.5
                top_fig = y0 + h
                cx_px, top_px = fig_to_pixel(cx_fig, top_fig)
                draw.text((cx_px, top_px), label, fill=fill, font=font, anchor="mb")

        if draw_scale_bar and data_width and data_width > 0 and last_ax_width_fig and scale_bar_px > 0:
            bar_length_px = scale_bar_px * (last_ax_width_fig / bw) * (W / data_width) if bw else 0
            bar_length_px = max(1, int(round(bar_length_px)))
            sx0, sy0 = scale_bar_pos.x0, scale_bar_pos.y0
            sw, sh = scale_bar_pos.width, scale_bar_pos.height
            cx_fig = sx0 + sw * 0.5
            cy_fig = sy0 + sh * 0.5
            cx_px, cy_px = fig_to_pixel(cx_fig, cy_fig)
            x1 = max(0, cx_px - bar_length_px // 2)
            x2 = min(W, cx_px + bar_length_px // 2)
            draw.line([(x1, cy_px), (x2, cy_px)], fill=fill, width=max(2, int(3 * dpi / 150)))
            label_lines = [f"{scale_bar_um:.0f} µm"]
            if scale_bar_caption:
                label_lines.append(scale_bar_caption)
            scale_bar_font_px = max(8, int(scale_bar_font_pt * dpi / 72))
            try:
                small_font = ImageFont.truetype("/System/Library/Fonts/Arial.ttf", scale_bar_font_px)
            except Exception:
                small_font = font
            draw.text((cx_px, cy_px + 12), "\n".join(label_lines), fill=fill, font=small_font, anchor="mt")

        # Element label: same layer, same pt→pixel scaling (so size matches other annotations)
        if draw_element_label and element_label_text:
            el_font_pt = max(6, min(72, element_label_font_pt))
            el_font_px = max(8, int(el_font_pt * dpi / 72))
            try:
                el_font = ImageFont.truetype("/System/Library/Fonts/Arial Bold.ttf", el_font_px)
            except Exception:
                try:
                    el_font = ImageFont.truetype("arialbd.ttf", el_font_px)
                except Exception:
                    el_font = font
            x_el, y_el = 50, H - 50
            draw.text((x_el, y_el), element_label_text, fill=fill, font=el_font, anchor="lb")

        # Optional credit / grant text in lower-right corner (subtle, not bright white)
        if draw_credit and credit_text:
            cr_font_pt = max(6, min(48, credit_font_pt))
            cr_font_px = max(6, int(cr_font_pt * dpi / 72))
            try:
                cr_font = ImageFont.truetype("/System/Library/Fonts/Arial.ttf", cr_font_px)
            except Exception:
                try:
                    cr_font = ImageFont.truetype("arial.ttf", cr_font_px)
                except Exception:
                    cr_font = font
            margin = max(10, int(12 * dpi / 150))
            x_cr = W - margin
            y_cr = H - margin
            # Subtle mid-gray so it doesn't compete with data
            credit_color = "#666666"
            try:
                draw.text((x_cr, y_cr), credit_text, fill=credit_color, font=cr_font, anchor="rd")
            except Exception:
                draw.text((x_cr - 5, y_cr - 5), credit_text, fill=credit_color, font=cr_font)

        return overlay

    def generate_composite(self, preview=False, override_rows=None):
        if not self.matrices:
            self.log_print("⚠️ No data loaded.")
            return

        # Ensure font dropdowns have committed their value to the variables (readonly Combobox updates on focus out)
        self.master.update_idletasks()
        self.set_status("Busy")
        if preview:
            self.log_print("Status: Busy - Generating preview...")
        else:
            self.log_print("Status: Busy - Generating composite...")

        # Auto-downsample when many samples (both preview and save)
        use_downsampling = len(self.matrices) > 10
        if use_downsampling:
            downsampled_matrices = [self.downsample_matrix(matrix) for matrix in self.matrices]
            matrices_to_use = downsampled_matrices
        else:
            matrices_to_use = self.matrices

        n = len(self.matrices)
        if override_rows is not None:
            rows = max(1, min(override_rows, n))
        elif getattr(self, 'use_best_layout', None) and self.use_best_layout.get():
            rows = self._best_composite_rows(n)
        else:
            rows = min(self.num_rows.get(), n)
        cols = math.ceil(n / rows)
        fig = plt.figure(figsize=(4 * cols + 1, 4 * rows))
        gs = fig.add_gridspec(rows, cols + 1, width_ratios=[1] * cols + [0.2])
        axs = np.empty((rows, cols + 1), dtype=object)
        for r in range(rows):
            for c in range(cols):
                axs[r, c] = fig.add_subplot(gs[r, c])
        for r in range(rows - 1):
            axs[r, cols] = fig.add_subplot(gs[r, cols])
        inner_gs = gs[rows - 1, cols].subgridspec(2, 1, height_ratios=[3.5, 1], hspace=0.15)
        color_bar_ax = fig.add_subplot(inner_gs[0, 0])
        scale_bar_ax = fig.add_subplot(inner_gs[1, 0])
        axs[rows - 1, cols] = color_bar_ax
        cmap = matplotlib.colormaps.get_cmap(self.color_scheme.get())

        scale_max = self.scale_max.get()

        if self.use_log.get():
            norm = self.pseudolog_norm(vmin=1, vmax=scale_max)
        else:
            norm = Normalize(vmin=0, vmax=scale_max)

        bg_color = cmap(0)
        fig.patch.set_facecolor(bg_color)
        text_color = 'white' if np.mean(bg_color[:3]) < 0.5 else 'black'

        percentiles = []
        iqrs = []
        means = []

        im = None
        show_subplot_label = str(self.sample_name_font.get()).strip() != "(None)"
        font_size = self._pt_from_font(self.sample_name_font, 12) if show_subplot_label else 12  # overlay + subtitles

        for i, (matrix, label) in enumerate(zip(matrices_to_use, self.labels)):
            if preview:
                self.log_print(f"   📊 Generating subplot {i+1}/{len(self.labels)}: {label}")
            r, c = i // cols, i % cols
            ax = axs[r, c]

            # Get the pixel size for this sample
            pixel_size = self.pixel_sizes_by_sample.get(label, self.pixel_size.get())
            H, W = matrix.shape[0], matrix.shape[1]
            if pixel_size and float(pixel_size) > 0:
                dx = float(pixel_size)
                extent = [0, W * dx, 0, H * dx]
                im = ax.imshow(matrix, cmap=cmap, norm=norm, aspect='equal', extent=extent)
                ax.set_aspect('equal')
            else:
                im = ax.imshow(matrix, cmap=cmap, norm=norm, aspect='auto')
                ax.set_aspect('auto')
            if show_subplot_label:
                # Sample names go on overlay layer (for future editing); keep title empty in base
                ax.set_title("", color=text_color, fontsize=font_size)
                if self.use_custom_pixel_sizes.get():
                    pixel_label = f"{int(round(pixel_size))} µm/px"
                    subtitle_size = (font_size - 2) if font_size else 9
                    subtitle_size = max(subtitle_size, 8)
                    ax.text(
                        0.5,
                        -0.08,
                        pixel_label,
                        transform=ax.transAxes,
                        ha='center',
                        va='top',
                        color=text_color,
                        fontsize=subtitle_size
                    )
            else:
                ax.set_title("", color=text_color)
            ax.axis('off')
            ax.set_facecolor(bg_color)

            # Save individual subplot (only if it doesn't exist - incremental processing)
            elem_out = self.get_element_output_subdir()
            subplot_path = os.path.join(self.output_dir, elem_out, 'subplots', f"{label}.png")
            os.makedirs(os.path.dirname(subplot_path), exist_ok=True)
            
            if not os.path.exists(subplot_path) or os.path.getsize(subplot_path) == 0:
                # Create a new figure for the individual subplot
                subplot_fig, subplot_ax = plt.subplots()
                subplot_fig.patch.set_facecolor(bg_color)
                subplot_ax.set_facecolor(bg_color)
                
                # Create a masked array for NaN values
                masked_matrix = np.ma.masked_where(np.isnan(matrix), matrix)
                if pixel_size and float(pixel_size) > 0:
                    dx = float(pixel_size)
                    extent = [0, W * dx, 0, H * dx]
                    subplot_ax.imshow(masked_matrix, cmap=cmap, norm=norm, aspect='equal', extent=extent)
                    subplot_ax.set_aspect('equal')
                else:
                    subplot_ax.imshow(masked_matrix, cmap=cmap, norm=norm, aspect='auto')
                    subplot_ax.set_aspect('auto')
                if show_subplot_label:
                    subplot_ax.set_title(f"{label}", color=text_color, fontsize=font_size)
                else:
                    subplot_ax.set_title("", color=text_color)
                subplot_ax.axis('off')
                subplot_fig.savefig(subplot_path, dpi=300, bbox_inches='tight', transparent=True)
                plt.close(subplot_fig)
                if preview:
                    # Subplot generated - progress table will show status
                    pass
            else:
                # Skipping - progress table will show status
                pass

            # Calculate percentiles, IQR, and mean
            p25, p50, p75, p99 = np.nanpercentile(matrix, [25, 50, 75, 99])
            iqr = p75 - p25
            mean = np.nanmean(matrix)
            percentiles.append((label, p25, p50, p75, p99))
            iqrs.append((label, iqr))
            means.append((label, mean))

        # Last image axes (for transform only; scale bar is drawn in right column)
        last_idx = len(matrices_to_use) - 1
        last_image_ax = axs[last_idx // cols, last_idx % cols]

        color_bar_ax.set_facecolor(bg_color)
        # Do not call axis('off') here so color bar tick labels (units) remain visible

        # Add color bar to its dedicated axes (top of right column)
        cbar_pt = self._pt_from_font(self.color_bar_font, 10)
        cbar = plt.colorbar(im, cax=color_bar_ax, orientation='vertical')
        cbar.ax.yaxis.set_tick_params(color=text_color, labelsize=cbar_pt)
        cbar.outline.set_edgecolor(text_color)
        plt.setp(plt.getp(cbar.ax.axes, 'yticklabels'), color=text_color)
        offset_text = cbar.ax.yaxis.get_offset_text()
        if offset_text:
            offset_text.set_color(text_color)
        # Units label above color bar (ppm, CPS, or counts)
        if getattr(self, 'current_element_unit', None):
            u = self.current_element_unit
            units_label = 'ppm' if u == 'ppm' else ('CPS' if u == 'CPS' else 'counts')
            cbar.set_label(units_label, color=text_color, fontsize=cbar_pt)

        # Scale bar in right column (below color bar): length from last image's data scale so it stays accurate
        if self.use_custom_pixel_sizes.get() and self.labels:
            # Only include the reference sample ID in the caption; omit per-sample pixel size text
            reference_label = self.labels[0]
            pixel_size_um = self.pixel_sizes_by_sample.get(reference_label, self.pixel_size.get())
            scale_bar_caption = reference_label
        else:
            reference_label = None
            pixel_size_um = self.pixel_size.get()
            scale_bar_caption = None

        scale_bar_um = self.scale_bar_length_um.get()
        if pixel_size_um <= 0:
            scale_bar_px = 0
        else:
            scale_bar_px = max(1, int(round(scale_bar_um / pixel_size_um)))

        # Scale bar and sample names drawn on overlay layer (same coords as base; locked together)
        scale_bar_ax.set_facecolor(bg_color)
        scale_bar_ax.axis('off')

        # Set background color and hide axes for image cells only (keep color bar labels visible)
        for ax in axs.flat:
            ax.set_facecolor(bg_color)
            if ax is not color_bar_ax:
                ax.axis('off')
        
        # Create standalone colorbar figure
        colorbar_fig, colorbar_ax = plt.subplots(figsize=(1, 4))
        colorbar_fig.patch.set_alpha(0.0)
        colorbar_ax.set_facecolor('none')

        # Re-create colorbar and apply styles again
        export_cbar = plt.colorbar(im, cax=colorbar_ax, orientation='vertical')

        # Reapply tick and outline styling (use same font size as main color bar)
        export_cbar.ax.yaxis.set_tick_params(color=text_color, labelsize=cbar_pt)
        export_cbar.outline.set_edgecolor(text_color)
        plt.setp(plt.getp(export_cbar.ax.axes, 'yticklabels'), color=text_color)
        # Set color for scientific notation offset label (e.g., "1e7")
        export_offset_text = export_cbar.ax.yaxis.get_offset_text()
        if export_offset_text:
            export_offset_text.set_color(text_color)
        if getattr(self, 'current_element_unit', None):
            u = self.current_element_unit
            units_label = 'ppm' if u == 'ppm' else ('CPS' if u == 'CPS' else 'counts')
            export_cbar.set_label(units_label, color=text_color, fontsize=cbar_pt)

        # Save
        elem_out = self.get_element_output_subdir()
        colorbar_path = os.path.join(self.output_dir, elem_out, f"{elem_out}_colorbar.png")
        colorbar_fig.savefig(colorbar_path, dpi=300, bbox_inches='tight', transparent=True)
        plt.close(colorbar_fig)

        # Overlay layer: sample names + scale bar (same coordinate system as base; locked together)
        # Save base without bbox_inches='tight' so figure coords (0-1) map 1:1 to pixels
        image_positions = [
            (axs[i // cols, i % cols].get_position().x0, axs[i // cols, i % cols].get_position().y0,
             axs[i // cols, i % cols].get_position().width, axs[i // cols, i % cols].get_position().height)
            for i in range(len(self.labels))
        ]
        scale_bar_pos = scale_bar_ax.get_position()
        last_ax_width_fig = last_image_ax.get_position().width
        dpi = 300
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi, facecolor=fig.get_facecolor())
        plt.close(fig)
        buf.seek(0)
        base_image = Image.open(buf).convert("RGB")
        # Element label: same overlay layer, same pt→pixel scaling
        el_val = self._get_element_label_font_value()
        draw_element_label = el_val != "(None)"
        element_label_text = ""
        if draw_element_label:
            element_name = self.element.get()
            units = "ppm"
            for file in glob.glob(os.path.join(self.input_dir or "", "* matrix.xlsx")):
                parsed = self.parse_matrix_filename(file)
                if parsed:
                    _, el, unit_type = parsed
                    if el == element_name:
                        units = "ppm" if unit_type == "ppm" else ("CPS" if unit_type == "CPS" else "counts")
                        break
            element_label_text = f"{element_name} ({units})"
        element_label_font_pt = self._pt_from_font_str(el_val, 16) if draw_element_label else 16
        # Optional credit / grant text in lower-right corner (opt-in)
        credit_text = (self.credit_text.get() or "").strip()
        draw_credit = bool(self.add_credit_to_exports.get() and credit_text)
        # Use full figure as bbox so (0,0)-(1,1) figure coords map to (0,0)-(W,H) pixels
        full_fig_bbox = type("Bbox", (), {"x0": 0, "y0": 0, "x1": 1, "y1": 1})()
        # Use aliases (if defined) for overlay labels; fall back to raw sample IDs
        overlay_labels = [self.sample_aliases.get(s, s) for s in self.labels]
        overlay = self._build_overlay_image(
            base_image,
            full_fig_bbox,
            image_positions,
            scale_bar_pos,
            last_ax_width_fig,
            overlay_labels,
            show_subplot_label,
            font_size if font_size is not None else 12,
            text_color,
            scale_bar_px,
            scale_bar_um,
            scale_bar_caption,
            matrices_to_use[-1].shape[1] if matrices_to_use else 0,
            draw_scale_bar=str(self.scale_bar_font.get()).strip() != "(None)",
            scale_bar_font_pt=self._pt_from_font(self.scale_bar_font, 10),
            draw_element_label=draw_element_label,
            element_label_text=element_label_text,
            element_label_font_pt=element_label_font_pt,
            draw_credit=draw_credit,
            credit_text=credit_text,
            credit_font_pt=8,
            dpi=dpi,
        )
        composited = Image.alpha_composite(base_image.convert("RGBA"), overlay).convert("RGB")

        # Store for future editing of sample names (redraw overlay only)
        self._base_preview_image = base_image
        self._overlay_sample_names = list(overlay_labels)

        if preview:
            self.preview_image = composited.copy()
            self.original_preview_image = self.preview_image.copy()
            self.preview_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False).name
            self.preview_image.save(self.preview_file)

            # Update preview in the Preview tab
            self.update_preview_image()

            # Switch to Preview tab to show the preview
            self.tabs.select(1)  # Index 1 is the Preview & Export tab
            self.set_status("Idle")
        else:
            elem_out = self.get_element_output_subdir()
            out_path = os.path.join(self.output_dir, elem_out, f"{elem_out}_composite.png")
            os.makedirs(os.path.dirname(out_path), exist_ok=True)
            composited.save(out_path)
            
            # Update progress table - mark as complete
            unit_type = self.unit.get() or 'ppm'
            for sample in self.labels:
                self.update_sample_element_progress(sample, self.element.get(), unit_type)
            
            # Update progress table display
            if hasattr(self, 'progress_table') and self.progress_table:
                self._check_existing_progress()
                self.update_progress_table()

            # Save percentiles, IQR, and mean table (including aliases)
            percentiles_df = pd.DataFrame(percentiles, columns=['Sample', '25th Percentile', '50th Percentile', '75th Percentile', '99th Percentile'])
            iqr_df = pd.DataFrame(iqrs, columns=['Sample', 'IQR'])
            mean_df = pd.DataFrame(means, columns=['Sample', 'Mean'])
            stats_df = percentiles_df.merge(iqr_df, on='Sample').merge(mean_df, on='Sample')
            stats_df['Alias'] = stats_df['Sample'].map(lambda s: self.sample_aliases.get(s, s))
            stats_path = os.path.join(self.output_dir, elem_out, f"{elem_out}_statistics.csv")
            stats_df.to_csv(stats_path, index=False)
            
            # Update statistics table display
            self.update_statistics_table(stats_df)
            
            self.set_status("Idle")
            self.log_print("Status: Idle - Composite saved.")

    def update_preview_image(self):
        """Update preview image in the Preview tab."""
        if not hasattr(self, 'preview_image') or self.preview_image is None:
            return
        
        # Force update to get accurate dimensions
        self.master.update_idletasks()
        self.preview_container.update_idletasks()
        
        container_width = self.preview_container.winfo_width()
        container_height = self.preview_container.winfo_height()
        
        # Handle zero or very small dimensions - use minimum sizes
        if container_width <= 10 or container_height <= 10:
            # Get window dimensions as fallback
            window_width = self.master.winfo_width()
            window_height = self.master.winfo_height()
            
            if window_width > 0 and window_height > 0:
                # Use most of the window space for preview, with minimums
                container_width = max(400, window_width - 250)  # Space for control panel
                container_height = max(300, window_height - 100)  # Space for window chrome
            else:
                # Default fallback dimensions - ensure minimum size
                container_width = max(400, 800)
                container_height = max(300, 600)
            
            # Force container to maintain minimum size
            self.preview_container.configure(width=max(400, container_width), height=max(300, container_height))
            
            # If we had to use fallback, schedule a retry to get proper dimensions
            if container_width == 800 and container_height == 600:
                self.master.after(100, self.update_preview_image)
                return
        
        img_width, img_height = self.preview_image.size
        
        # Basic safety check for image dimensions
        if img_width <= 0 or img_height <= 0:
            self.log_print(f"⚠️ Invalid image dimensions: {img_width}x{img_height}")
            return
        
        # Calculate aspect ratios
        img_aspect = img_width / img_height
        container_aspect = container_width / container_height
        
        # Resize to fit container while maintaining aspect ratio
        if container_aspect > img_aspect:
            # Container is wider - fit to height
            new_height = container_height
            new_width = int(new_height * img_aspect)
        else:
            # Container is taller - fit to width
            new_width = container_width
            new_height = int(new_width / img_aspect)
        
        # Ensure minimum size
        if new_width < 10 or new_height < 10:
            self.log_print(f"⚠️ Calculated size too small: {new_width}x{new_height}, using fallback")
            new_width = max(400, container_width)
            new_height = max(300, container_height)
        
        try:
            resized_image = self.preview_image.resize((new_width, new_height), Image.LANCZOS)
            tk_image = ImageTk.PhotoImage(resized_image)
            
            self.preview_label.config(image=tk_image)
            self.preview_label.image = tk_image  # Keep a reference to prevent garbage collection
        except Exception as e:
            # Log the error but don't crash the application
            self.log_print(f"⚠️ Warning: Failed to resize preview image: {e}")
            import traceback
            self.log_print(traceback.format_exc())

    def _add_element_label_to_image(self, img):
        """Draw element label on a PIL image if Element label font is not (None). Returns new image (or unchanged if off)."""
        el_val = self._get_element_label_font_value()
        if not img or el_val == "(None)":
            return img.copy() if img else img
        try:
            from PIL import ImageDraw, ImageFont
            labeled_image = img.convert("RGB").copy()
            draw = ImageDraw.Draw(labeled_image)
            element_name = self.element.get()
            units = "ppm"
            for file in glob.glob(os.path.join(self.input_dir or "", "* matrix.xlsx")):
                parsed = self.parse_matrix_filename(file)
                if parsed:
                    _, element, unit_type = parsed
                    if element == element_name:
                        units = "ppm" if unit_type == "ppm" else ("CPS" if unit_type == "CPS" else "counts")
                        break
            img_width, img_height = labeled_image.size
            font_size = self._pt_from_font_str(el_val, 16)
            font_size = max(6, font_size)  # only enforce a small minimum for readability
            try:
                font = ImageFont.truetype("/System/Library/Fonts/Arial Bold.ttf", font_size)
            except Exception:
                try:
                    font = ImageFont.truetype("arialbd.ttf", font_size)
                except Exception:
                    try:
                        font = ImageFont.truetype("arial.ttf", font_size)
                    except Exception:
                        font = ImageFont.load_default()
            element_text = f"{element_name} ({units})"
            x, y = 50, img_height - 50
            draw.text((x, y), element_text, fill=(255, 255, 255), font=font, anchor="lb")
            return labeled_image
        except Exception:
            return img.copy() if img else img

    def add_element_label(self):
        """Element label is drawn in the overlay; changing the dropdown triggers full refresh via trace. This runs refresh so preview updates (e.g. if called from elsewhere)."""
        self._on_font_change_refresh_preview()

def main():
    root = tk.Tk()
    app = CompositeApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()