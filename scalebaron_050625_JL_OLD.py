"""
ScaleBarOn v0.8.6 — GitHub-Ready Composite Generator
Date: 2025-05-08

CHANGELOG:
- Formal versioning system initiated (v0.8.0)
- Base version derived from `scalebaron_clean2`
- Reorganized and documented all code blocks
- Preserved all core features:
  - Pixel size prompt
  - Color scheme and layout prompts
  - Image rotation (optional)
  - File inventory confirmation
  - Scale bar size prompt (µm)
  - Composite organization (grid, spatial and ppm scale bars)
  - Summary tables (99th percentile, variance)
- v0.8.1: Improved efficiency by using subplots instead of saving/reloading images
- v0.8.2: Set background of plots to lowest value in color scale, adjust title text color for visibility
- v0.8.3: Prompt for number of rows, automatically calculate columns, adjust colorbar placement
- v0.8.4: Set unused plots in final column to background color, narrow colorbar, add pseudo-log scale option
- v0.8.5: Fixed colorbar visibility issue, ensured all unused axes have background set and are turned off
- v0.8.6: Adjusted colorbar and scale bar text color based on background brightness

Next goals:
- Further optimize visual layout of scale bars and labels
- Track cumulative features for JOSS submission
"""

import os
import re
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib import cm
from scipy import stats
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
import glob

def main():
    # --- User Inputs ---
    PIXEL_SIZE = float(input("Enter pixel size in microns: ") or 6) 
    SCALE_BAR_LENGTH_UM = float(input("Enter desired scale bar length in microns: ") or 500) 
    SCALE_BAR_PIXELS = int(SCALE_BAR_LENGTH_UM // PIXEL_SIZE)

    color_scheme = input("Enter a Matplotlib color scheme (e.g., jet, viridis): ") or "jet"
    rotate_images = input("Rotate images? (y/n): ").strip().lower() == 'y'
    element = input("Enter element name: ") or "Cu63"
    use_pseudo_log = input("Use pseudo-log scale for concentration values? (y/n): ").strip().lower() == 'y'
    portrait = None
    if rotate_images:
        layout_mode = input("Composite orientation: portrait or landscape? (p/l): ").strip().lower()
        portrait = layout_mode == 'p'

    INPUT_DIR = "./testdata"
    OUTPUT_DIR = "./OUTPUT"
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # --- Utility Functions ---
    def load_matrix_2d(path):
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active
        return np.array([[cell if isinstance(cell, (int, float)) else np.nan for cell in row] for row in ws.iter_rows(values_only=True)])

    def compute_99th_percentile(values):
        return np.nanpercentile(values.flatten(), 99)

    def compute_variance(values):
        return np.nanvar(values.flatten())

    def pseudo_log_scale(x, a=1):
        return np.sign(x) * np.log1p(np.abs(x) / a)

    def create_composite(matrices, labels, scale_max, unit, element):
        total_images = len(matrices)
        rows = int(input("Enter number of rows for layout: "))
        cols = math.ceil(total_images / rows)
        
        fig, axs = plt.subplots(rows, cols + 1, figsize=(4 * (cols + 1), 4 * rows))
        cmap = plt.get_cmap(color_scheme)
        
        if use_pseudo_log:
            norm = mcolors.FuncNorm((lambda x: pseudo_log_scale(x), lambda x: np.expm1(x)), vmin=0, vmax=scale_max)
        else:
            norm = mcolors.Normalize(vmin=0, vmax=scale_max)

        # Set background color to lowest value in color scale
        bg_color = cmap(0)
        fig.patch.set_facecolor(bg_color)

        # Determine text color based on background brightness
        bg_brightness = np.mean(bg_color[:3])
        text_color = 'white' if bg_brightness < 0.5 else 'black'

        for i, (matrix, label) in enumerate(zip(matrices, labels)):
            r, c = i // cols, i % cols
            ax = axs[r, c] if rows > 1 else axs[c]
            
            # Use the original matrix without filling NaN values
            im = ax.imshow(matrix, cmap=cmap, norm=norm)
            
            # Set background color for the subplot
            ax.set_facecolor(bg_color)
            
            ax.set_title(label, color=text_color)
            ax.axis('off')

        # Add colorbar in the middle row of the last column
        middle_row = rows // 2
        outer_ax = axs[middle_row, -1] if rows > 1 else axs[-1]
        outer_ax.set_facecolor(bg_color)
        outer_ax.axis('off')
        cax = inset_axes(outer_ax,
                 width="30%",    # 30% of the width of the subplot
                 height="100%",  # full height
                 loc='center',   # center the inset in the subplot
                 bbox_to_anchor=(0, 0, 1, 1),  # occupy full space of outer_ax
                 bbox_transform=outer_ax.transAxes,
                 borderpad=0)

        cbar = plt.colorbar(im, cax=cax, orientation='vertical', label=unit)
        
        # Set background color for unused colorbar area
        cbar.outline.set_visible(True)  # Ensure the outline is visible
        cax.set_facecolor(bg_color)  # Set background to lowest color
        
        # Add min and max values to colorbar with appropriate text color
        cbar.ax.text(0, 0, f'{0:.2f}', ha='right', va='bottom', transform=cbar.ax.transAxes, color=text_color)
        cbar.ax.text(0, 1, f'{scale_max:.2f}', ha='right', va='top', transform=cbar.ax.transAxes, color=text_color)
        cbar.ax.yaxis.set_tick_params(color=text_color)
        cbar.ax.yaxis.label.set_color(text_color)
        plt.setp(plt.getp(cbar.ax.axes, 'yticklabels'), color=text_color)

        # Add scale bar
        scale_bar_ax = axs[0, -1] if rows > 1 else axs[-1]
        scale_bar_ax.set_xlim(0, 1)
        scale_bar_ax.set_ylim(0, 1)
        scale_bar_ax.add_line(plt.Line2D([0.1, 0.1 + SCALE_BAR_PIXELS/100], [0.1, 0.1], color=text_color, linewidth=2))
        scale_bar_ax.text(0.1, 0.15, f"{int(SCALE_BAR_LENGTH_UM)} µm", color=text_color, ha='left', va='bottom')
        scale_bar_ax.axis('off')
        scale_bar_ax.set_facecolor(bg_color)

        # Set unused plots in the final column to background color and turn off axes
        for i in range(total_images, rows * cols):
            r, c = i // cols, i % cols
            ax = axs[r, c] if rows > 1 else axs[c]
            ax.set_facecolor(bg_color)
            ax.axis('off')

        # Ensure all unused axes in the last column are set to background color and turned off
        for r in range(rows):
            if r != middle_row:  # Skip the row with the colorbar
                ax = axs[r, -1] if rows > 1 else axs[-1]
                ax.set_facecolor(bg_color)
                ax.axis('off')

        plt.tight_layout()
        out_path = os.path.join(OUTPUT_DIR, element, f"{element}_composite.png")
        plt.savefig(out_path, dpi=300, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close(fig)

    # --- File Inventory ---
    print("\n🔍 Scanning INPUT folder...")
    matrix_files = sorted(os.path.basename(f) for f in glob.glob(os.path.join(INPUT_DIR, f"* {element}_ppm matrix.xlsx")))
    print(f"Found {len(matrix_files)} files:")
    for f in matrix_files:
        print(" •", f)

    if input("Proceed? (y/n): ").lower() != 'y':
        print("Aborted.")
        return

    # --- Main Processing ---
    samples_by_element = {}
    for file in matrix_files:
        match = re.match(r"(.+)[ _]([A-Za-z]{1,2}\d{2,3})_(ppm|CPS) matrix\.xlsx", file)
        if match:
            sample, element, unit_type = match.groups()
            key = f"{element}_{unit_type}"
            samples_by_element.setdefault(key, []).append((sample, os.path.join(INPUT_DIR, file)))
        else:
            print(f"⚠️ Skipped: {file}")

    summary_rows = []
    variance_rows = []

    for key, sample_list in samples_by_element.items():
        element, unit_type = key.split("_")
        unit = "ppm" if unit_type == "ppm" else "CPS"
        element_dir = os.path.join(OUTPUT_DIR, element)
        os.makedirs(element_dir, exist_ok=True)

        print(f"\n📊 Processing {len(sample_list)} sample(s) for {element} [{unit}]")

        all_values = []
        sample_99th = {}
        sample_var = {}
        matrices = []
        labels = []

        for sample, filepath in sample_list:
            matrix = load_matrix_2d(filepath)
            all_values.append(matrix)
            sample_99th[sample] = compute_99th_percentile(matrix)
            sample_var[sample] = compute_variance(matrix)
            print(f"   • {sample}: 99th = {sample_99th[sample]:.2f}, variance = {sample_var[sample]:.2f}")
            matrices.append(matrix)
            labels.append(sample)

        flat_all = np.concatenate([m.flatten() for m in all_values])
        scale_max = compute_99th_percentile(flat_all)

        create_composite(matrices, labels, scale_max, unit, element)

        for sample in sample_99th:
            summary_rows.append({"Sample": sample, element: round(sample_99th[sample], 1)})
            variance_rows.append({"Sample": sample, element: round(sample_var[sample], 1)})

    # --- Save Summary Tables ---
    if summary_rows:
        df = pd.DataFrame(summary_rows).groupby("Sample").first().reset_index()
        var_df = pd.DataFrame(variance_rows).groupby("Sample").first().reset_index()
        df.to_csv(os.path.join(OUTPUT_DIR, "summary_99th_percentiles.csv"), index=False)
        var_df.to_csv(os.path.join(OUTPUT_DIR, "summary_variance.csv"), index=False)
        print("\n✅ Summary tables saved to OUTPUT")

    print("\n✅ All processing complete.")

if __name__ == "__main__":
    main()
