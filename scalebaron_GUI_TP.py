# This script sets up a GUI application version of your command-line tool using tkinter
# Core features: Load data, choose element, layout options, preview, and save composite

import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.colors import Normalize
from matplotlib import cm
from openpyxl import load_workbook
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
import math
import glob
import re
import tempfile
from PIL import Image, ImageTk
import shutil
import pandas as pd

class CompositeApp:
    def __init__(self, master):
        self.master = master
        master.title("ScaleBarOn v0.8.6")

        self.pixel_size = tk.DoubleVar(value=6.0)
        self.use_multiple_pixel_sizes = tk.BooleanVar(value=False)
        self.pixel_sizes_by_sample = {}  # Dictionary to store per-sample pixel sizes
        self.scale_bar_length_um = tk.DoubleVar(value=500.0)
        self.num_rows = tk.IntVar(value=5)
        self.use_log = tk.BooleanVar(value=False)
        self.color_scheme = tk.StringVar(value="jet")
        self.rotate = tk.BooleanVar(value=False)
        self.element = tk.StringVar()
        self.sample_name_font_size = tk.StringVar(value="n/a")  # Default to "n/a"
        self.scale_max = tk.DoubleVar(value=1.0)  # New variable for scale_max

        self.input_dir = None
        self.output_dir = "./OUTPUT"
        os.makedirs(self.output_dir, exist_ok=True)

        self.matrices = []
        self.labels = []
        self.preview_file = None

        self.setup_widgets()

    def toggle_pixel_size_mode(self):
        if self.use_multiple_pixel_sizes.get():
            self.metadata_button.state(['!disabled'])
        else:
            self.metadata_button.state(['disabled'])
            self.pixel_sizes_by_sample = {}  # Clear mapping

    def load_metadata_file(self):
        file_path = filedialog.askopenfilename(filetypes=[('CSV files', '*.csv')])
        if not file_path:
            return
        try:
            df = pd.read_csv(file_path)
            self.pixel_sizes_by_sample = dict(zip(df['Sample'].astype(str), df['PixelSize']))
            self.log_print(f'✅ Loaded pixel size metadata for {len(self.pixel_sizes_by_sample)} samples.')
        except Exception as e:
            messagebox.showerror('Error', f'Failed to load metadata file:\n{e}')
            self.pixel_sizes_by_sample = {}
    def setup_widgets(self):
        control_frame = ttk.Frame(self.master)
        control_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)

        preview_frame = ttk.Frame(self.master)
        preview_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Controls
        ttk.Button(control_frame, text="Select Input Folder", command=self.select_input_folder).pack(pady=5)

        grid_frame = ttk.Frame(control_frame)
        grid_frame.pack(pady=10)

        ttk.Label(grid_frame, text="Element:").grid(row=0, column=0, sticky="e", padx=5, pady=2)
        self.element_dropdown = ttk.Combobox(grid_frame, textvariable=self.element, state="disabled")
        self.element_dropdown.grid(row=0, column=1, padx=5, pady=2)

        ttk.Label(grid_frame, text="Pixel size (µm):").grid(row=1, column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(grid_frame, textvariable=self.pixel_size).grid(row=1, column=1, padx=5, pady=2)
        self.multiple_pixel_check = ttk.Checkbutton(grid_frame, text='Use Multiple Pixel Sizes', variable=self.use_multiple_pixel_sizes, command=self.toggle_pixel_size_mode)
        self.multiple_pixel_check.grid(row=1, column=2, padx=5, pady=2)
        self.metadata_button = ttk.Button(grid_frame, text='Select Metadata File', command=self.load_metadata_file)
        self.metadata_button.grid(row=2, column=2, padx=5, pady=2)
        self.metadata_button.state(['disabled'])

        ttk.Label(grid_frame, text="Scale bar length (µm):").grid(row=2, column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(grid_frame, textvariable=self.scale_bar_length_um).grid(row=2, column=1, padx=5, pady=2)

        ttk.Label(grid_frame, text="Rows:").grid(row=3, column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(grid_frame, textvariable=self.num_rows).grid(row=3, column=1, padx=5, pady=2)

        ttk.Label(grid_frame, text="Color Scheme:").grid(row=4, column=0, sticky="e", padx=5, pady=2)
        self.color_scheme_dropdown = ttk.Combobox(grid_frame, textvariable=self.color_scheme, values=plt.colormaps())
        self.color_scheme_dropdown.grid(row=4, column=1, padx=5, pady=2)

        ttk.Label(grid_frame, text="Sample Name Font Size:").grid(row=5, column=0, sticky="e", padx=5, pady=2)
        self.sample_name_font_size_dropdown = ttk.Combobox(grid_frame, textvariable=self.sample_name_font_size, values=["n/a", "Small", "Medium", "Large"])
        self.sample_name_font_size_dropdown.grid(row=5, column=1, padx=5, pady=2)

        ttk.Label(grid_frame, text="Scale Max:").grid(row=6, column=0, sticky="e", padx=5, pady=2)
        ttk.Entry(grid_frame, textvariable=self.scale_max).grid(row=6, column=1, padx=5, pady=2)

        ttk.Checkbutton(grid_frame, text="Use Pseudo-log", variable=self.use_log).grid(row=7, column=0, columnspan=2, pady=2)
        ttk.Checkbutton(grid_frame, text="Rotate Images", variable=self.rotate).grid(row=8, column=0, columnspan=2, pady=2)

        button_frame = ttk.Frame(control_frame)
        button_frame.pack(pady=10)

        ttk.Button(button_frame, text="Load Data", command=self.load_data).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(button_frame, text="Preview", command=self.preview_composite).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(button_frame, text="Select Output Folder", command=self.select_output_folder).grid(row=1, column=0, padx=5, pady=5)
        ttk.Button(button_frame, text="Save", command=self.save_composite).grid(row=1, column=1, padx=5, pady=5)

        self.log = tk.Text(control_frame, height=20, width=40)
        self.log.pack(pady=10)

        # Preview frame
        self.preview_container = ttk.Frame(preview_frame)
        self.preview_container.pack(fill=tk.BOTH, expand=True)
        
        self.preview_label = ttk.Label(self.preview_container)
        self.preview_label.pack(fill=tk.BOTH, expand=True)
        
        # Bind resize event to update the preview image's aspect ratio
        self.preview_container.bind("<Configure>", self.on_resize)

    def select_input_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.input_dir = folder_selected
            self.log_print(f"Input folder updated to: {self.input_dir}")
            self.update_element_dropdown()

    def select_output_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.output_dir = folder_selected
            self.log_print(f"Output folder updated to: {self.output_dir}")

    def update_element_dropdown(self):
        elements = set()
        for file in glob.glob(os.path.join(self.input_dir, "*.xlsx")):
            match = re.search(r"([A-Za-z]{1,2}\d{2,3})_ppm matrix\.xlsx", file)
            if match:
                elements.add(match.group(1))
        
        if elements:
            self.element_dropdown['values'] = sorted(list(elements))
            self.element_dropdown['state'] = 'readonly'
            self.element.set(next(iter(elements)))  # Set the first element as default
        else:
            self.log_print("No valid element files found in the selected directory.")

    def on_resize(self, event):
        if hasattr(self, 'preview_image'):
            self.update_preview_image()

    def log_print(self, message):
        self.log.insert(tk.END, message + "\n")
        self.log.see(tk.END)
        self.master.update_idletasks()  # Force update of the GUI

    def load_matrix_2d(self, path):
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active
        return np.array([[cell if isinstance(cell, (int, float)) else np.nan for cell in row] for row in ws.iter_rows(values_only=True)])

    def load_data(self):
        if not self.input_dir:
            messagebox.showerror("Error", "Please select an input folder first.")
            return

        self.log_print("\n🔍 Scanning INPUT folder...")
        element = self.element.get()
        pattern = os.path.join(self.input_dir, f"* {element}_ppm matrix.xlsx")
        files = sorted(glob.glob(pattern))

        if not files:
            messagebox.showerror("Error", f"No files found for element {element}")
            return

        self.matrices = []
        self.labels = []
        percentiles = []
        iqrs = []

        for f in files:
            match = re.match(r"(.+)[ _]([A-Za-z]{1,2}\d{2,3})_(ppm|CPS) matrix\.xlsx", os.path.basename(f))
            if match:
                sample, _, _ = match.groups()
                self.labels.append(sample)
                matrix = self.load_matrix_2d(f)
                self.matrices.append(matrix)
                
                # Calculate percentiles and IQR
                p25, p50, p75, p99 = np.nanpercentile(matrix, [25, 50, 75, 99])
                iqr = p75 - p25
                percentiles.append((sample, p25, p50, p75, p99))
                iqrs.append((sample, iqr))
                
                self.log_print(f"Loaded: {sample}")
                self.log_print(f"  99th percentile: {p99:.2f}")
                self.log_print(f"  IQR: {iqr:.2f}")
                
                # Generate and save histogram
                plt.figure(figsize=(10, 6))
                plt.hist(matrix.flatten(), bins=50, range=(0, np.nanpercentile(matrix, 99)))
                plt.title(f"Histogram for {sample}")
                plt.xlabel("Value")
                plt.ylabel("Frequency")
                hist_path = os.path.join(self.output_dir, self.element.get(), 'histograms', f"{sample}_histogram.png")
                os.makedirs(os.path.dirname(hist_path), exist_ok=True)
                plt.savefig(hist_path)
                plt.close()

        self.log_print(f"✅ Loaded {len(self.matrices)} matrix files.")
        self.log_print(f"Histograms saved in: {os.path.join(self.output_dir, self.element.get(), 'histograms')}")

        # Save percentiles and IQR table
        percentiles_df = pd.DataFrame(percentiles, columns=['Sample', '25th Percentile', '50th Percentile', '75th Percentile', '99th Percentile'])
        iqr_df = pd.DataFrame(iqrs, columns=['Sample', 'IQR'])
        stats_df = percentiles_df.merge(iqr_df, on='Sample')
        stats_path = os.path.join(self.output_dir, self.element.get(), f"{self.element.get()}_statistics.csv")
        stats_df.to_csv(stats_path, index=False)
        self.log_print(f"✅ Saved statistics table to {stats_path}")

        # Set scale_max based on 99th percentile of all data
        overall_99th = np.nanpercentile(np.hstack([m.flatten() for m in self.matrices]), 99)
        self.scale_max.set(overall_99th)
        self.log_print(f"Scale max set to {self.scale_max.get():.2f} based on overall 99th percentile")

    def preview_composite(self):
        self.generate_composite(preview=True)

    def save_composite(self):
        if self.preview_file:
            out_path = os.path.join(self.output_dir, self.element.get(), f"{self.element.get()}_composite.png")
            os.makedirs(os.path.dirname(out_path), exist_ok=True)
            shutil.move(self.preview_file, out_path)
            self.log_print(f"✅ Saved composite to {out_path}")
            self.preview_file = None
        else:
            self.generate_composite(preview=False)

    def generate_composite(self, preview=False):
        if not self.matrices:
            self.log_print("⚠️ No data loaded.")
            return

        rows = self.num_rows.get()
        cols = math.ceil(len(self.matrices) / rows)
        fig, axs = plt.subplots(rows, cols + 1, figsize=(4 * cols + 1, 4 * rows), gridspec_kw={'width_ratios': [1] * cols + [0.2]})
        cmap = cm.get_cmap(self.color_scheme.get())

        scale_max = self.scale_max.get()

        if self.use_log.get():
            norm = Normalize(vmin=0, vmax=scale_max)
        else:
            norm = Normalize(vmin=0, vmax=scale_max)

        bg_color = cmap(0)
        fig.patch.set_facecolor(bg_color)
        text_color = 'white' if np.mean(bg_color[:3]) < 0.5 else 'black'

        # Find the maximum dimensions of all matrices
        max_height = max(matrix.shape[0] for matrix in self.matrices)
        max_width = max(matrix.shape[1] for matrix in self.matrices)

        percentiles = []
        iqrs = []

        for i, (matrix, label) in enumerate(zip(self.matrices, self.labels)):
            r, c = i // cols, i % cols
            ax = axs[r, c] if rows > 1 else axs[c]

            # Calculate padding to center the matrix
            pad_height = (max_height - matrix.shape[0]) // 2
            pad_width = (max_width - matrix.shape[1]) // 2

            # Pad the matrix to match the maximum dimensions and center it
            padded_matrix = np.pad(matrix, 
                                   ((pad_height, max_height - matrix.shape[0] - pad_height), 
                                    (pad_width, max_width - matrix.shape[1] - pad_width)), 
                                   mode='constant', constant_values=np.nan)
            
            # Determine font size based on selection
            font_size = None
            if self.sample_name_font_size.get() == "Small":
                font_size = 8
            elif self.sample_name_font_size.get() == "Medium":
                font_size = 12
            elif self.sample_name_font_size.get() == "Large":
                font_size = 16

            im = ax.imshow(padded_matrix, cmap=cmap, norm=norm, aspect='equal')
            ax.set_title(label, color=text_color, fontsize=font_size)
            ax.axis('off')
            ax.set_facecolor(bg_color)

            # Save individual subplot
            if not preview:
                subplot_path = os.path.join(self.output_dir, self.element.get(), 'subplots', f"{label}.png")
                os.makedirs(os.path.dirname(subplot_path), exist_ok=True)
                extent = ax.get_window_extent().transformed(fig.dpi_scale_trans.inverted())
                fig.savefig(subplot_path, bbox_inches=extent.expanded(1.1, 1.2), dpi=300)

            # Calculate percentiles and IQR
            p25, p50, p75, p99 = np.nanpercentile(matrix, [25, 50, 75, 99])
            iqr = p75 - p25
            percentiles.append((label, p25, p50, p75, p99))
            iqrs.append((label, iqr))

        # Set color to bg_color and drop axes for the final column of subplots
        for r in range(rows):
            ax = axs[r, -1] if rows > 1 else axs[-1]
            ax.set_facecolor(bg_color)
            ax.axis('off')

        # Add color bar and scale bar to the last column
        last_ax = axs[-1, -1] if rows > 1 else axs[-1]
        last_ax.axis('off')
        last_ax.set_facecolor(bg_color)

        # Create two inset axes for color bar and scale bar
        color_ax = inset_axes(last_ax, width="50%", height="60%", loc='upper left',
                              bbox_to_anchor=(0, 1, 1, 0.6), bbox_transform=last_ax.transAxes)
        scale_ax = inset_axes(last_ax, width="100%", height="40%", loc='lower left',
                              bbox_to_anchor=(0, 0, 1, 0.4), bbox_transform=last_ax.transAxes)

        # Add color bar
        cbar = plt.colorbar(im, cax=color_ax, orientation='vertical')
        cbar.ax.yaxis.set_tick_params(color=text_color)
        cbar.outline.set_edgecolor(text_color)
        plt.setp(plt.getp(cbar.ax.axes, 'yticklabels'), color=text_color)

        # Add scale bar
        pixel_size = self.pixel_sizes_by_sample.get(label, self.pixel_size.get())
        scale_bar_length_pixels = self.scale_bar_length_um.get() / pixel_size
        scale_bar_width = scale_bar_length_pixels / max_width
        scale_ax.add_line(plt.Line2D([0.1, 0.1 + scale_bar_width], [0.5, 0.5], color=text_color, linewidth=2, transform=scale_ax.transAxes))
        scale_ax.text(0.1, 0.7, f"{int(self.scale_bar_length_um.get())} µm", color=text_color, ha='left', va='bottom', transform=scale_ax.transAxes)
        scale_ax.axis('off')

        plt.tight_layout()
        
        if preview:
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                plt.savefig(tmp_file.name, dpi=300, bbox_inches='tight', facecolor=fig.get_facecolor())
            plt.close(fig)
            
            self.preview_image = Image.open(tmp_file.name)
            self.update_preview_image()
            
            self.preview_file = tmp_file.name
            self.log_print("Preview generated. Click 'Save' to keep this image.")
        else:
            out_path = os.path.join(self.output_dir, self.element.get(), f"{self.element.get()}_composite.png")
            os.makedirs(os.path.dirname(out_path), exist_ok=True)
            plt.savefig(out_path, dpi=300, bbox_inches='tight', facecolor=fig.get_facecolor())
            plt.close(fig)
            self.log_print(f"✅ Saved composite to {out_path}")

            # Save percentiles and IQR table
            percentiles_df = pd.DataFrame(percentiles, columns=['Sample', '25th Percentile', '50th Percentile', '75th Percentile', '99th Percentile'])
            iqr_df = pd.DataFrame(iqrs, columns=['Sample', 'IQR'])
            stats_df = percentiles_df.merge(iqr_df, on='Sample')
            stats_path = os.path.join(self.output_dir, self.element.get(), f"{self.element.get()}_statistics.csv")
            stats_df.to_csv(stats_path, index=False)
            self.log_print(f"✅ Saved statistics table to {stats_path}")

    def update_preview_image(self):
        if hasattr(self, 'preview_image'):
            container_width = self.preview_container.winfo_width()
            container_height = self.preview_container.winfo_height()
            
            img_width, img_height = self.preview_image.size
            aspect_ratio = img_width / img_height
            
            if container_width / container_height > aspect_ratio:
                new_height = container_height
                new_width = int(new_height * aspect_ratio)
            else:
                new_width = container_width
                new_height = int(new_width / aspect_ratio)
            
            resized_image = self.preview_image.resize((new_width, new_height), Image.LANCZOS)
            tk_image = ImageTk.PhotoImage(resized_image)
            
            self.preview_label.config(image=tk_image)
            self.preview_label.image = tk_image  # Keep a reference to prevent garbage collection

if __name__ == '__main__':
    root = tk.Tk()
    app = CompositeApp(root)
    root.mainloop()
